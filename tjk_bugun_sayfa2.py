#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
TJK - sadece BUGUN Sayfa1 + Sayfa2 uretir.

Bu dosya gun_ayar.py icindeki HEDEF_TARIH/OFFSET ayarini kullanmaz.
Amaci: bugunun kosu programindan Sayfa1 ve Sayfa2 olusturup ayri bir Excel'e yazmak.
"""

import os
import time
from datetime import datetime

from openpyxl import Workbook
from openpyxl.styles import Font

import tjk_yeni_yer as YY
import tjk_donustur as DON


CIKTI_DOSYA = "BUGUN_SAYFA1_SAYFA2.xlsx"


def bugun_programi_cek():
    from selenium.webdriver.common.by import By

    hedef_tarih = datetime.now().strftime("%d/%m/%Y")
    driver = YY._driver_baslat()
    sehir_sayfa1 = []
    try:
        driver.get(YY.ANA_URL)
        time.sleep(2)
        print("AKTIF URL:", driver.current_url)
        YY._hedef_gune_git(driver, hedef_tarih)

        sehir_link = {}
        for a in driver.find_elements(By.TAG_NAME, "a"):
            try:
                href = a.get_attribute("href") or ""
                txt = (a.text or "").strip()
                if "SehirId=" not in href or "GunlukYarisProgrami" not in href:
                    continue
                for s in YY.SEHIRLER:
                    if s.lower() in txt.lower() and s not in sehir_link:
                        sehir_link[s] = href
            except Exception:
                continue

        print("Bulunan sehirler:", list(sehir_link.keys()))
        if not sehir_link:
            print("UYARI: Bugun Turkiye sehri bulunamadi.")

        for sehir, url in sehir_link.items():
            print(f"\n--- {sehir} ---")
            try:
                driver.get(url)
                time.sleep(2)
                try:
                    driver.execute_script("window.scrollTo(0, document.body.scrollHeight);")
                    time.sleep(1)
                    driver.execute_script("window.scrollTo(0, 0);")
                    time.sleep(1)
                except Exception:
                    pass
                html = driver.page_source
            except Exception as e:
                print(f"  {sehir} acilamadi: {e}")
                continue

            satirlar, kosu_say = YY._html_to_sayfa1(html, 0)
            if satirlar:
                sehir_sayfa1.append((sehir, satirlar))
            print(f"  {sehir}: {kosu_say} kosu, {len(satirlar)} ham satir")
    finally:
        try:
            driver.quit()
        except Exception:
            pass

    print(f"\nToplam il: {len(sehir_sayfa1)}")
    return sehir_sayfa1


def yaz(sayfa1_all, sayfa2_all):
    wb = Workbook()

    ws1 = wb.active
    ws1.title = "Sayfa1"
    for row in sayfa1_all:
        ws1.append(["" if v is None else v for v in row])

    ws2 = wb.create_sheet("Sayfa2")
    DON.yaz_sayfa2(ws2, sayfa2_all)

    for ws in (ws1, ws2):
        for cell in ws[1]:
            cell.font = Font(bold=True)

    wb.save(CIKTI_DOSYA)


def main():
    print("=" * 60)
    print("TJK - BUGUN Sayfa1 + Sayfa2")
    print("=" * 60)

    il_listesi = bugun_programi_cek()
    if not il_listesi:
        raise SystemExit("Bugun program cekilemedi.")

    sayfa1_all = []
    sayfa2_all = []
    for sehir, sayfa1_city in il_listesi:
        if sayfa1_all:
            sayfa1_all.append([f"--- {sehir} ---"] + [None] * 11)
        sayfa1_all.extend(sayfa1_city)
        s2 = DON.sayfa1_to_sayfa2_rows(sayfa1_city)
        sayfa2_all.extend(s2)
        print(f"{sehir}: Sayfa2 {len(s2)} satir")

    yaz(sayfa1_all, sayfa2_all)
    print(f"\nTAMAM: {os.path.abspath(CIKTI_DOSYA)}")


if __name__ == "__main__":
    main()
