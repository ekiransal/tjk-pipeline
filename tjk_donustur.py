#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
TJK günlük yarış programı dönüştürücü
=====================================

Bu betik, orijinal Excel makrosunun ("TJK_Hamdan_Makroya_10lu") yaptığı işin
birebir Python karşılığıdır.

Ne yapar:
  - Girdi Excel'inin "Sayfa1" sayfasındaki kazınmış (scrape edilmiş) TJK
    verisini okur.
  - Her koşu bloğunu ve o bloktaki at satırlarını bulur.
  - Her atı 10 kez çoğaltır (Rep = 1..10) ve aşağıdaki kolonları üretir.
  - Koşu bazında Handikap Puanı (HP) ve Kilo (Siklet) "tam ortanca"
    (medyan) değerlerini hesaplayıp her satıra yazar.
  - Sonucu yeni bir Excel dosyasının "Sayfa2" sayfasına yazar.

Çıktı kolonları (A..L):
  Kosu_No, At_No, At_Adi_Temiz, Siklet, Rep, Anahtar, Kosu_Basligi,
  Yas, KGS, Handikap_Puani, Kosu_HP_Tam_Ortanca, Kosu_Kilo_Tam_Ortanca

Kullanım:
  python3 tjk_donustur.py girdi.xlsm                 -> girdi_cikti.xlsx
  python3 tjk_donustur.py girdi.xlsm cikti.xlsx
"""

import re
import sys
import ast
import operator
import datetime
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font


# ----------------------------------------------------------------------------
# Yardımcı fonksiyonlar (VBA fonksiyonlarının karşılıkları)
# ----------------------------------------------------------------------------

# Atın isminin sonundaki "kuyruk kodları" (KG, DB, SK, vb.). Temizlikte atılır.
TAIL_CODES = {
    "KG", "KUL", "DB", "BB", "ÇABA",
    "GKR", "SKG", "SKUL", "SGKR",
    "YP", "TGK", "ÖG", "DS",
    "SK", "K", "T",
}


def is_tail_code(token: str) -> bool:
    """VBA IsTailCode: bir parça (token) at adının sonundan atılacak kod mu?"""
    token = (token or "").strip().upper()
    for ch in (".", ",", ")", "("):
        token = token.replace(ch, "")
    if len(token) == 0:
        return True
    if len(token) == 1:          # tek harf/karakter -> kuyruk kabul edilir
        return True
    return token in TAIL_CODES


def clean_horse_name(s) -> str:
    """VBA CleanHorseName: at isminden '(Satılık)', '%' ve sondaki kuyruk
    kodlarını temizler."""
    s = "" if s is None else str(s).strip()
    # '(Satılık)' ifadesini büyük/küçük harf duyarsız sil
    s = re.sub(r"\(Satılık\)", "", s, flags=re.IGNORECASE)
    s = s.replace("%", " ")
    if len(s) == 0:
        return ""
    # WorksheetFunction.Trim: birden çok boşluğu teke indir + baş/son boşluk
    s = re.sub(r"\s+", " ", s).strip()
    # Web programından gelen kirli kuyruğu kes: ödül tutarı (t1.500.000,00 TL)
    # ve ekipman notları (KGKapalı gözlük..., DBDilinin..., SKRing...).
    # TJK at adları TAMAMEN BÜYÜK harf; kuyruk küçük harf/rakam içerir.
    # İlk küçük-harf veya rakam içeren token'da kes.
    _toks = s.split(" ")
    _kept = []
    for _t in _toks:
        if re.search(r"[a-zçğıiöşü0-9]", _t):
            break
        _kept.append(_t)
    if _kept:
        s = " ".join(_kept)
    parts = s.split(" ") if s else []
    # Sondan başlayarak kuyruk kodlarını at
    i = len(parts) - 1
    while i >= 0:
        if is_tail_code(parts[i]):
            if i == 0:
                return ""
            parts = parts[:i]
            i -= 1
        else:
            break
    return " ".join(parts).strip()


# Güvenli aritmetik değerlendirme (Application.Evaluate yerine).
# Sadece + - * / ( ) ve sayılar; örn. "54+2" -> 56. Excel'deki gibi.
_ALLOWED_BINOPS = {
    ast.Add: operator.add,
    ast.Sub: operator.sub,
    ast.Mult: operator.mul,
    ast.Div: operator.truediv,
}
_ALLOWED_UNARY = {ast.UAdd: operator.pos, ast.USub: operator.neg}


def _safe_eval_arith(expr: str):
    node = ast.parse(expr, mode="eval").body

    def ev(n):
        if isinstance(n, ast.Constant) and isinstance(n.value, (int, float)):
            return n.value
        if isinstance(n, ast.BinOp) and type(n.op) in _ALLOWED_BINOPS:
            return _ALLOWED_BINOPS[type(n.op)](ev(n.left), ev(n.right))
        if isinstance(n, ast.UnaryOp) and type(n.op) in _ALLOWED_UNARY:
            return _ALLOWED_UNARY[type(n.op)](ev(n.operand))
        raise ValueError("izin verilmeyen ifade")

    return ev(node)


def sum_plus_weight(v):
    """VBA SumPlusWeight: sıklet hücresini değerlendirir.
    - boş -> ""
    - sayı -> float
    - '54+2' gibi ifade -> hesaplanmış değer
    - hatalı -> ""
    """
    if v is None or str(v).strip() == "":
        return ""
    if isinstance(v, (int, float)) and not isinstance(v, bool):
        return float(v)
    s = str(v).strip().replace(" ", "").replace(",", ".")
    try:
        total = _safe_eval_arith(s)
    except Exception:
        return ""
    if total is None or str(total) == "":
        return ""
    return total


def is_race_title(s) -> bool:
    """VBA IsRaceTitle: 'Koşu:' içeriyor mu ya da '#. Koşu' / '##. Koşu'
    kalıbıyla başlıyor mu?"""
    s = "" if s is None else str(s).strip()
    if "koşu:" in s.lower():
        return True
    return re.match(r"^\d{1,2}\. Koşu", s) is not None


def race_no_from_title(race_title: str) -> str:
    """VBA RaceNoFromTitle: başlıktaki ilk '.'ten önceki koşu numarası."""
    s = (race_title or "").strip()
    p = s.find(".")
    if p > 0:  # VBA InStr>1 demek: en az 1 karakter solda olmalı (p>0, 0-index)
        return s[:p].strip()
    return ""


def _to_long(v):
    """VBA CLng benzeri: yarım yukarı değil, banker's rounding'e yakın değil;
    Excel CLng 'round half to even' kullanır. At numaraları tam sayı olduğu
    için pratikte fark etmez, ama yine de yakın davranış."""
    if v is None:
        raise ValueError
    if isinstance(v, bool):
        raise ValueError
    if isinstance(v, (int, float)):
        return int(round(v))
    s = str(v).strip().replace(",", ".")
    # tarih/metin -> hata
    f = float(s)
    return int(round(f))


def is_horse_row(ws, r: int) -> bool:
    """VBA IsHorseRow: A sütununda 1..99 arası numara, B'de boş olmayan
    ve 'At İsmi' olmayan bir ad varsa at satırıdır."""
    v_no = ws.cell(r, 1).value
    v_name = ws.cell(r, 2).value
    v_name = "" if v_name is None else str(v_name)
    if len(str("" if v_no is None else v_no).strip()) == 0:
        return False
    if len(v_name.strip()) == 0:
        return False
    if v_name.strip().upper() == "AT İSMİ":
        return False
    # tarih ise at satırı değil
    if isinstance(v_no, (datetime.datetime, datetime.date)):
        return False
    try:
        n = _to_long(v_no)
    except Exception:
        return False
    if n < 1 or n > 99:
        return False
    return True


def median_of(values):
    """VBA MedianOfArray: klasik medyan (çift sayıda elemanda ortalama)."""
    if not values:
        return ""
    arr = sorted(values)
    n = len(arr)
    if n % 2 == 1:
        return arr[(n - 1) // 2]          # (cnt+1)\2, 1-index -> 0-index
    return (arr[n // 2 - 1] + arr[n // 2]) / 2.0


def _is_number(v):
    if v is None or isinstance(v, bool):
        return False
    if isinstance(v, (int, float)):
        return True
    try:
        float(str(v).strip().replace(",", "."))
        return True
    except Exception:
        return False


def _as_float(v):
    if isinstance(v, (int, float)):
        return float(v)
    return float(str(v).strip().replace(",", "."))


# ----------------------------------------------------------------------------
# Ana dönüştürme
# ----------------------------------------------------------------------------

class _RowsWS:
    """Bellek-içi satır listesini openpyxl worksheet gibi gösteren shim.
    rows: list[list]  (rows[r-1][c-1] = hücre).  1-tabanlı cell(r,c) erişimi."""
    class _Cell:
        __slots__ = ("value",)
        def __init__(self, value):
            self.value = value
    def __init__(self, rows):
        self._rows = rows
        self.max_row = len(rows)
        self.max_column = max((len(r) for r in rows), default=0)
    def cell(self, r, c):
        if 1 <= r <= len(self._rows):
            row = self._rows[r - 1]
            if 1 <= c <= len(row):
                return _RowsWS._Cell(row[c - 1])
        return _RowsWS._Cell(None)


def sayfa1_to_sayfa2_rows(sayfa1_rows):
    """Bellek-içi: Sayfa1 satır listesi -> Sayfa2 veri satırları (12 kolon, başlıksız).
    sayfa1_rows: list[list], 1-tabanlı kolonlar 0-index'te (rows[i][0]=A)."""
    return _build_sayfa2_rows(_RowsWS(sayfa1_rows))


# Sayfa2 başlıkları (modül genelinde tek kaynak)
SAYFA2_HEADERS = [
    "Kosu_No", "At_No", "At_Adi_Temiz", "Siklet", "Rep", "Anahtar",
    "Kosu_Basligi", "Yas", "KGS", "Handikap_Puani",
    "Kosu_HP_Tam_Ortanca", "Kosu_Kilo_Tam_Ortanca",
]


def _build_sayfa2_rows(ws):
    """Verilen (ws benzeri) Sayfa1 kaynağından Sayfa2 veri satırlarını üretir."""
    # Son dolu satır (VBA Cells.Find xlPrevious karşılığı)
    last_row = 0
    for r in range(ws.max_row, 0, -1):
        if any(ws.cell(r, c).value not in (None, "") for c in range(1, ws.max_column + 1)):
            last_row = r
            break
    if last_row == 0:
        raise SystemExit("Sayfa1 boş.")

    # Çıktı satırları: her biri 12 kolonluk liste (A..L)
    rows_out = []

    r = 1
    race_title = ""
    race_no = ""
    while r <= last_row:
        a_val = ws.cell(r, 1).value
        b_val = ws.cell(r, 2).value
        a_str = "" if a_val is None else str(a_val)
        b_str = "" if b_val is None else str(b_val)

        if is_race_title(a_val):
            race_title = a_str.strip()
            race_no = race_no_from_title(race_title)
            r += 1
        elif a_str == "N" and b_str == "At İsmi":
            # At tablosu başlığı -> at satırlarını işle
            r += 1
            while r <= last_row:
                a2 = ws.cell(r, 1).value
                b2 = ws.cell(r, 2).value
                a2s = "" if a2 is None else str(a2)
                b2s = "" if b2 is None else str(b2)
                if is_race_title(a2):
                    break
                if a2s == "N" and b2s == "At İsmi":
                    break
                if is_horse_row(ws, r):
                    # Kosu_No tamamen rakamsa Excel'in makroda yaptığı gibi
                    # sayı olarak sakla (aksi halde metin kalır).
                    race_no_cell = int(race_no) if race_no.isdigit() else race_no
                    at_no = _to_long(ws.cell(r, 1).value)
                    at_adi_temiz = clean_horse_name(ws.cell(r, 2).value)
                    siklet_val = sum_plus_weight(ws.cell(r, 5).value)
                    yas_val = "" if ws.cell(r, 3).value is None else str(ws.cell(r, 3).value).strip()
                    hp_val = ws.cell(r, 10).value
                    kgs_val = ws.cell(r, 12).value
                    for rep in range(1, 11):
                        rows_out.append([
                            race_no_cell,                  # 1  Kosu_No
                            at_no,                         # 2  At_No
                            at_adi_temiz,                  # 3  At_Adi_Temiz
                            siklet_val,                    # 4  Siklet
                            rep,                           # 5  Rep
                            f"{at_adi_temiz}{rep}",        # 6  Anahtar
                            race_title,                    # 7  Kosu_Basligi
                            yas_val,                       # 8  Yas
                            kgs_val,                       # 9  KGS
                            hp_val,                        # 10 Handikap_Puani
                            "",                            # 11 HP medyan (sonra)
                            "",                            # 12 Kilo medyan (sonra)
                        ])
                r += 1
        else:
            r += 1

    # ---- Koşu bazında tam ortanca (medyan) kolonlarını doldur ----
    # VBA mantığı: her koşu için Rep==1 olan satırlardan HP'yi (kol 10) topla;
    # HP sayısal ise say. Kilo (kol 4) yalnızca ilgili HP sayısalken eklenir.
    medians = {}  # race_no -> (hp_medyan, kilo_medyan)
    for row in rows_out:
        rn = str(row[0]).strip()
        if rn == "" or rn in medians:
            continue
        hp_vals, kilo_vals = [], []
        for rr in rows_out:
            if str(rr[0]).strip() == rn and _as_float(rr[4]) == 1:  # Rep == 1
                if str(rr[9]).strip() != "" and _is_number(rr[9]):
                    hp_vals.append(_as_float(rr[9]))
                    if str(rr[3]).strip() != "" and _is_number(rr[3]):
                        kilo_vals.append(_as_float(rr[3]))
        medians[rn] = (median_of(hp_vals), median_of(kilo_vals))

    for row in rows_out:
        rn = str(row[0]).strip()
        if rn in medians:
            row[10], row[11] = medians[rn]

    return rows_out


def yaz_sayfa2(ws_out, rows_out):
    """Verilen openpyxl worksheet'ine Sayfa2 başlığı + satırlarını yazar."""
    ws_out.append(SAYFA2_HEADERS)
    for c in range(1, 13):
        ws_out.cell(1, c).font = Font(bold=True)
    for row in rows_out:
        ws_out.append(["" if v is None else v for v in row])
    for col in (4, 11, 12):
        for rr in range(2, ws_out.max_row + 1):
            ws_out.cell(rr, col).number_format = "0.##"


def donustur(input_path: str, output_path: str):
    wb_in = load_workbook(input_path, data_only=True)
    if "Sayfa1" not in wb_in.sheetnames:
        raise SystemExit("Girdi dosyasında 'Sayfa1' sayfası yok.")
    rows_out = _build_sayfa2_rows(wb_in["Sayfa1"])
    wb_out = Workbook()
    ws_out = wb_out.active
    ws_out.title = "Sayfa2"
    yaz_sayfa2(ws_out, rows_out)
    wb_out.save(output_path)
    return len(rows_out), output_path


def main():
    if len(sys.argv) < 2:
        print("Kullanım: python3 tjk_donustur.py girdi.xlsm [cikti.xlsx]")
        raise SystemExit(1)
    inp = sys.argv[1]
    if len(sys.argv) >= 3:
        out = sys.argv[2]
    else:
        out = re.sub(r"\.(xlsm|xlsx|xls)$", "", inp, flags=re.IGNORECASE) + "_cikti.xlsx"
    n, path = donustur(inp, out)
    print(f"İşlem tamamlandı. Üretilen satır sayısı: {n}")
    print(f"Çıktı dosyası: {path}")


if __name__ == "__main__":
    main()
