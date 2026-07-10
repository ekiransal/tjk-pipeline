#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""sayfa internet excell -> mockup_parsed.json (koşu kartları + detay satırları)."""
import openpyxl, json, re, sys

KAYNAK = sys.argv[1] if len(sys.argv) > 1 else "sayfa_internet.xlsx"
TABLO_AD = {"KALİTE": "Kalite", "MESAFE": "Mesafe", "SPRİNTER": "Sprinter", "KAÇAK": "Kaçak"}
GALOP_RE = re.compile(r"(1000\+400|800\+400|600\+400|400\s*fark)", re.I)
TARIH_RE = re.compile(r"^\d{2}\.\d{2}\.\d{4}$")


def _tarih(s):
    return bool(TARIH_RE.match(str(s).strip()))


def parse_sheet(ws):
    grid = {}
    for r in range(1, ws.max_row + 1):
        for c in range(1, ws.max_column + 1):
            v = ws.cell(r, c).value
            if v is not None and str(v).strip() != "":
                grid[(r, c)] = v
    maxr, maxc = ws.max_row, ws.max_column
    starts = [r for r in range(1, maxr + 1) if str(grid.get((r, 1), "")).strip() == "Koşu No"]
    bloklar = []
    GALV = re.compile(r"^\d+\s*/")                                   # galop hücresi: "9 / -4,60 / HÇ"
    GALO = re.compile(r"^\d+\s*/\s*[-+−]?\d+[.,]\d+\s*/")            # kesin galop deseni (öksüz satır avı)
    for bi, rs in enumerate(starts):
        re_ = starts[bi + 1] - 1 if bi + 1 < len(starts) else maxr
        # blok içi başlık satırı ("N. Koşu ...") -> galop taraması bunun ÜSTÜNDE kalmalı (detaya girmesin)
        sinir = re_
        for r0 in range(rs + 1, re_ + 1):
            if re.match(r"^\d+\.\s*Koşu", str(grid.get((r0, 1), "")).strip()):
                sinir = r0 - 1
                break
        blok = {"header": {}, "tables": [], "galops": [], "detay": [], "title": ""}
        c = 1
        while c <= maxc:
            lab = grid.get((rs, c))
            if lab is not None:
                val = grid.get((rs, c + 1))
                blok["header"][str(lab).strip()] = "" if val is None else str(val).strip()
                c += 2
            else:
                c += 1
        for r in range(rs + 1, re_ + 1):
            rowv = {c: str(grid.get((r, c), "")).strip() for c in range(1, maxc + 1) if (r, c) in grid}
            names = {c: v for c, v in rowv.items() if v.upper() in TABLO_AD}
            if names and any(str(grid.get((r + 1, c), "")).strip() == "No" for c in names):
                for c0, nm in sorted(names.items()):
                    rows = []
                    rr = r + 2
                    while rr <= re_:
                        no = grid.get((rr, c0)); deg = grid.get((rr, c0 + 1)); say = grid.get((rr, c0 + 2))
                        if no is None and deg is None:
                            break
                        # KORUMA: orjin satırı 'No=tam sayı, Değer=sayı' olmalı; galop/şehir
                        # satırı orjinin altına yapışmışsa tablo BURADA biter (Kocaeli 8 vakası).
                        _no = str(no).strip() if no is not None else ""
                        try:
                            float(str(deg).replace(",", "."))
                            _degok = True
                        except Exception:
                            _degok = False
                        if not (re.match(r"^\d+$", _no) and _degok):
                            break
                        rows.append([("" if x is None else str(x)) for x in (no, deg, say)])
                        rr += 1
                    blok["tables"].append({"name": TABLO_AD[nm.upper()], "col": c0, "rows": rows})
            gal = {c: v for c, v in rowv.items() if GALOP_RE.search(v)}
            if gal:
                for c0, nm in sorted(gal.items()):
                    isSon = str(nm).strip().upper().startswith("SON")
                    rows = []
                    rr = r + 2
                    bos = 0
                    while rr <= sinir:                       # detaya asla sarkmaz
                        v = grid.get((rr, c0))
                        s = "" if v is None else str(v).strip()
                        if s == "":
                            bos += 1                          # BOŞLUK TOLERANSI: veri birkaç satır
                            if bos > 8:                       # aşağıdan başlayabiliyor (bozuk yerleşim)
                                break
                            rr += 1
                            continue
                        if not GALV.match(s):                 # galop deseni değilse bitti
                            break
                        bos = 0
                        t = grid.get((rr, c0 + 1))
                        sh = grid.get((rr, c0 + 2)) if isSon else None
                        rows.append([s,
                                     "" if t is None else str(t).strip(),
                                     "" if sh is None else str(sh).strip()])
                        rr += 1
                    blok["galops"].append({"name": str(nm).strip(), "son": isSon, "col": c0, "rows": rows})
            v1 = rowv.get(1, "")
            if re.match(r"^\d+\.\s*Koşu", v1):
                blok["title"] = v1
            # DETAY SATIRI (üç düzen de desteklenir):
            def _metin(s):
                s = str(s).strip()
                if s == "" or s in ("Değer", "No", "Sayı"):
                    return False
                try:
                    float(s.replace(",", ".")); return False
                except Exception:
                    return any(ch.isalpha() for ch in s)
            c1v, c2v = rowv.get(1, ""), rowv.get(2, "")
            eski_birlesik = (_metin(c2v) and _tarih(rowv.get(3, ""))
                             and _metin(rowv.get(21, "")) and _tarih(rowv.get(28, "")))
            eski_devam = (not c1v and _metin(rowv.get(21, "")) and _tarih(rowv.get(28, "")))
            yeni_duzen = (c1v and re.match(r"^\d+([.,]\d+)?$", c1v) and _metin(c2v)
                          and (_tarih(rowv.get(3, "")) or _tarih(rowv.get(9, ""))))
            if eski_birlesik or eski_devam:
                # ESKİ birleşik düzen (800 solda): derece kısmı eski 20-54 -> 1-35 kaydırılır
                blok["detay"].append(
                    ["" if grid.get((r, c)) is None else str(grid.get((r, c))) for c in range(20, 55)])
            elif yeni_duzen:
                # YENİ v7 düzeni (sola yaslı) ya da 800 sayfası: olduğu gibi
                blok["detay"].append(
                    ["" if grid.get((r, c)) is None else str(grid.get((r, c))) for c in range(1, 43)])
        # ÖKSÜZ GALOP KURTARMA: bozuk yerleşimde ilk galop satırları başlıksız olarak
        # SOL kolonlara (1,3,5 / 8,11,14) yazılıyor. Desen: hücre=galop, sağı=tarih.
        # Bu gruplar soldan sağa, panellerin sırasını AYNEN izler -> sırayla eşle,
        # kurtarılan satırlar panelin BAŞINA eklenir. Sayılar tutmazsa DOKUNMA.
        if blok["galops"]:
            kapsanan = set()
            for g in blok["galops"]:
                kapsanan.update(range(g["col"], g["col"] + 3))
            orf = {}
            for r2 in range(rs + 1, sinir + 1):
                for c2 in range(1, maxc + 1):
                    if c2 in kapsanan or (r2, c2) not in grid:
                        continue
                    s = str(grid[(r2, c2)]).strip()
                    if GALO.match(s) and _tarih(str(grid.get((r2, c2 + 1), "")).strip()):
                        orf.setdefault(c2, []).append(r2)
            if orf:
                def _sonmu(c2):
                    # şehir hücresi: harfli, '/'siz, tarih değil ("Ankara"); yandaki grubun
                    # galop hücresi ('5 / -2,30 / R') şehir SAYILMAZ.
                    for r2 in orf[c2]:
                        sh = str(grid.get((r2, c2 + 2), "")).strip()
                        if sh and "/" not in sh and not _tarih(sh) and any(ch.isalpha() for ch in sh):
                            return True
                    return False
                normO = sorted(c2 for c2 in orf if not _sonmu(c2))
                sonO = sorted(c2 for c2 in orf if _sonmu(c2))
                normP = [g for g in blok["galops"] if not g["son"]]
                sonP = [g for g in blok["galops"] if g["son"]]
                kurtarilan = 0
                if len(normO) == len(normP):
                    for c2, g in zip(normO, normP):
                        ek = [[str(grid[(r2, c2)]).strip(),
                               str(grid.get((r2, c2 + 1), "")).strip(), ""] for r2 in sorted(orf[c2])]
                        g["rows"] = ek + g["rows"]; kurtarilan += len(ek)
                if len(sonO) == len(sonP):
                    for c2, g in zip(sonO, sonP):
                        ek = [[str(grid[(r2, c2)]).strip(),
                               str(grid.get((r2, c2 + 1), "")).strip(),
                               str(grid.get((r2, c2 + 2), "")).strip()] for r2 in sorted(orf[c2])]
                        g["rows"] = ek + g["rows"]; kurtarilan += len(ek)
                if kurtarilan:
                    print(f"  [KURTARMA] {blok['header'].get('İl','?')} {blok['header'].get('Koşu No','?')}. koşu: "
                          f"{kurtarilan} öksüz galop satırı yerine kondu")
        bloklar.append(blok)
    return bloklar


wb = openpyxl.load_workbook(KAYNAK, data_only=True)

# Sheet eşlemesi: gerçek pipeline çıktısı (yeni_yer_SONUC.xlsx) ya da elle mockup.
#   Sayfa1 rolü = Toplam Derece, Sayfa2 rolü = Son 800.
if "yapılacak yer" in wb.sheetnames:
    esleme = {"Sayfa1": "yapılacak yer", "Sayfa2": "yapılacak yer 800"}
else:
    esleme = {"Sayfa1": "Sayfa1", "Sayfa2": "Sayfa2"}

out = {}
for rol, sh in esleme.items():
    if sh in wb.sheetnames:
        out[rol] = parse_sheet(wb[sh])
        det = sum(len(b["detay"]) for b in out[rol])
        print(f"{rol} <- '{sh}': {len(out[rol])} koşu | detay satırı: {det}")

# ---------------------------------------------------------------------------
# KAYIP AT KORUMASI (İSEN BUGA vakası):
#   Derece verisi 'son 6 ayda İLK 4' filtreli olduğundan hiç ilk-4 yapamayan at
#   TD tablosunda GÖRÜNMEZ oluyordu. Koşunun at listesi (800 sayfası) ile TD
#   karşılaştırılır; eksik atlara işaretli tek satır eklenir -> sitede
#   "son 6 ayda ilk-4 derecesi bulunmuyor" diye görünür. At kaybolmaz.
# ---------------------------------------------------------------------------
if "Sayfa1" in out and "Sayfa2" in out:
    def _norm(s):
        return re.sub(r"\s+", " ", str(s or "").strip()).upper()
    # 1) ASIL kaynak: ham program listesi (workbook'taki 'Sayfa2' sayfası:
    #    Kosu_No / At_No / At_Adi_Temiz / Kosu_Basligi) -> koşu başlığına göre eşlenir.
    program = {}   # norm(başlık) -> [(at_no, at_adi), ...]
    try:
        if "Sayfa2" in wb.sheetnames:
            wsp = wb["Sayfa2"]
            hp = [str(c.value or "").strip() for c in wsp[1]]
            if "At_Adi_Temiz" in hp and "Kosu_Basligi" in hp:
                iN, iA, iB = hp.index("At_No"), hp.index("At_Adi_Temiz"), hp.index("Kosu_Basligi")
                for row in wsp.iter_rows(min_row=2, values_only=True):
                    at = str(row[iA] or "").strip().upper()
                    no = str(row[iN] or "").strip()
                    bas = _norm(row[iB])
                    if at and bas:
                        L = program.setdefault(bas, [])
                        if (no, at) not in L:
                            L.append((no, at))
    except Exception as _e:
        print(f"  NOT: ham program listesi okunamadı ({_e}); 800 listesi kullanılacak")
    _eklenen = 0
    for b1 in out["Sayfa1"]:
        il = b1["header"].get("İl", ""); kosu = b1["header"].get("Koşu No", "")
        # asıl liste: program (başlıkla); yoksa 800 sayfası (yedek)
        liste = program.get(_norm(b1.get("title", "")))
        if not liste:
            b2 = next((b for b in out["Sayfa2"]
                       if b["header"].get("İl") == il and str(b["header"].get("Koşu No")) == str(kosu)), None)
            liste = []
            if b2:
                g = set()
                for r in b2["detay"]:
                    at = str(r[1]).strip().upper(); no = str(r[0]).strip()
                    if at and at not in g:
                        g.add(at); liste.append((no, at))
        var_olan = {re.sub(r"\d+$", "", str(r[1]).strip().upper()).strip() for r in b1["detay"]}
        for no, at in liste:
            if at and at not in var_olan:
                yeni = [""] * 42
                yeni[0] = no; yeni[1] = at; yeni[41] = "GECMIS_YOK"
                b1["detay"].append(yeni)
                var_olan.add(at)
                _eklenen += 1
    if _eklenen:
        print(f"Kayıp at koruması: {_eklenen} at TD'ye 'derecesi yok' satırıyla eklendi")

# ---------------------------------------------------------------------------
# SON 800 DOMİNANS HİZALAMA (BAY OLOF vakası):
#   Pipeline 800 sayfasında dominans satır kaymasıyla YANLIŞ gelebiliyor
#   (koşu anahtarı tutmayınca bir üst satırın değerine düşüyor).
#   DOĞRUSU: aynı at+tarih+şehir koşusunun Toplam Derece dominansı.
#   TD'de olmayan koşuların dominansı BOŞ bırakılır — yanlış göstermekten iyidir.
# ---------------------------------------------------------------------------
if "Sayfa1" in out and "Sayfa2" in out:
    td_map = {}
    for b in out["Sayfa1"]:
        for r in b["detay"]:
            at = re.sub(r"\d+$", "", str(r[1]).strip().upper()).strip()
            if not at:
                continue
            key = (at, str(r[8]).strip(), str(r[9]).strip().upper())
            td_map[key] = [(r[i] if i < len(r) else "") for i in range(24, 30)]
    _duz = _bos = 0
    for b in out["Sayfa2"]:
        for r in b["detay"]:
            at = str(r[1]).strip().upper()
            if not at:
                continue
            key = (at, str(r[2]).strip(), str(r[3]).strip().upper())
            dom = td_map.get(key)
            while len(r) < 25:
                r.append("")
            if dom:
                for i in range(6):
                    r[19 + i] = dom[i]
                _duz += 1
            else:
                _bos += 1   # TD penceresi dışı (6ay-375g): pipeline'ın (yamalı) kendi değeri KORUNUR
    print(f"800 dominans hizalama: {_duz} satır TD'den eşlendi, {_bos} satır TD dışı (pipeline değeri korundu)")

# ---------------------------------------------------------------------------
# EXTREMLER: uzun aradan gelen + çok sık koşan atlar (sağ üst kutu için)
# ---------------------------------------------------------------------------
import datetime
# IRKA GÖRE EŞİKLER (jokey görüşü: İngiliz dinlenmeye muhtaç, yakın koşması dikkat çeker;
# Arap dayanıklı, sık koşabilir AMA uzun aradan daha çok etkilenir)
SIK_ARAP,  SIK_ING  = 5, 6      # son koşusu <= bu kadar gün önce -> "sık koşan"
UZUN_ARAP, UZUN_ING = 45, 75    # son koşusu >= bu kadar gün önce -> "uzun ara"
def _esik(bas):
    """Koşu başlığından ırkı bulur, (sık_eşik, uzun_eşik) döndürür. Arap değilse İngiliz varsayılır."""
    return (SIK_ARAP, UZUN_ARAP) if "ARAP" in str(bas or "").upper() else (SIK_ING, UZUN_ING)


def _hedef_tarih():
    for yol in ("../gun_ayar.py", "gun_ayar.py"):
        try:
            s = open(yol, encoding="utf-8").read()
            m = re.search(r'(?m)^HEDEF_TARIH\s*=\s*"(\d{2}\.\d{2}\.\d{4})"', s)
            if m:
                return datetime.datetime.strptime(m.group(1), "%d.%m.%Y").date()
        except Exception:
            pass
    return None


def _d(s):
    try:
        return datetime.datetime.strptime(str(s).strip(), "%d.%m.%Y").date()
    except Exception:
        return None


hedef = _hedef_tarih()

# -------- YENİ YÖNTEM: TJK'nın RESMİ 'KGS' kolonu (koşmadığı gün sayısı) --------
# Eski yöntem tablolardaki son koşu tarihinden hesaplıyordu; tablolar filtreli
# (ilk-4 / 375 gün) olduğundan gerçek son koşu kaçıyor, gün sayısı ŞİŞİYORDU
# (350g görünen at gerçekte 20g önce koşmuş olabiliyordu). KGS resmî ve kesindir.
extrem = []
_kgs_ok = False
try:
    if "Sayfa2" in wb.sheetnames:
        wsp = wb["Sayfa2"]
        hp = [str(c.value or "").strip() for c in wsp[1]]
        if "KGS" in hp and "At_Adi_Temiz" in hp and "Kosu_Basligi" in hp:
            iN, iA, iK, iB = (hp.index("At_No"), hp.index("At_Adi_Temiz"),
                              hp.index("KGS"), hp.index("Kosu_Basligi"))
            def _norm2(s):
                return re.sub(r"\s+", " ", str(s or "").strip()).upper()
            baslik_il = {}
            for b in out.get("Sayfa1", []):
                baslik_il[_norm2(b.get("title", ""))] = (b["header"].get("İl", ""),
                                                         b["header"].get("Koşu No", ""))
            gorulen = set()
            for row in wsp.iter_rows(min_row=2, values_only=True):
                at = str(row[iA] or "").strip().upper()
                bas = _norm2(row[iB])
                if not at or bas not in baslik_il or (bas, at) in gorulen:
                    continue
                gorulen.add((bas, at))
                try:
                    gun = int(float(str(row[iK]).replace(",", ".")))
                except Exception:
                    continue                       # KGS boş (ilk koşusu) -> extrem'e girmez
                il, kosu = baslik_il[bas]
                _sik, _uzun = _esik(bas)   # bas = koşu başlığı (ırk bilgisi içinde)
                if gun <= _sik:
                    extrem.append({"il": il, "kosu": kosu, "atno": str(row[iN] or "").strip(),
                                   "at": at, "gun": gun, "tip": "sik"})
                elif gun >= _uzun:
                    extrem.append({"il": il, "kosu": kosu, "atno": str(row[iN] or "").strip(),
                                   "at": at, "gun": gun, "tip": "uzun"})
            _kgs_ok = True
            print("Extremler: resmi KGS kolonundan hesaplandı")
except Exception as _e:
    print(f"  NOT: KGS okunamadı ({_e}); eski tarih yöntemine dönülüyor")

if not _kgs_ok:
    son_kosu = {}   # (il, kosu, at) -> {"atno":..., "son": date}
    tum = []
    for rol, ad_i, tar_i, sufix in (("Sayfa2", 1, 2, False), ("Sayfa1", 1, 8, True)):
        for b in out.get(rol, []):
            il = b["header"].get("İl", ""); kosu = b["header"].get("Koşu No", "")
            _bas = b.get("title", "")   # ırk bilgisi başlıkta
            for r in b["detay"]:
                at = str(r[ad_i]).strip()
                if sufix:
                    at = re.sub(r"\d+$", "", at).strip()
                t = _d(r[tar_i])
                if not at or t is None:
                    continue
                tum.append(t)
                k = (il, kosu, at.upper())
                if k not in son_kosu or t > son_kosu[k]["son"]:
                    son_kosu[k] = {"atno": str(r[0]).strip(), "son": t, "bas": _bas}
    if hedef is None and tum:
        hedef = max(tum) + datetime.timedelta(days=1)
    if hedef:
        for (il, kosu, at), v in son_kosu.items():
            gun = (hedef - v["son"]).days
            _sik, _uzun = _esik(v.get("bas", ""))
            if gun <= _sik:
                extrem.append({"il": il, "kosu": kosu, "atno": v["atno"], "at": at,
                               "gun": gun, "tip": "sik"})
            elif gun >= _uzun:
                extrem.append({"il": il, "kosu": kosu, "atno": v["atno"], "at": at,
                               "gun": gun, "tip": "uzun"})
extrem.sort(key=lambda x: (x["il"], x["tip"], -x["gun"]))
out["extremler"] = {"hedef": hedef.strftime("%d.%m.%Y") if hedef else "",
                    "sik_arap": SIK_ARAP, "sik_ing": SIK_ING,
                    "uzun_arap": UZUN_ARAP, "uzun_ing": UZUN_ING, "liste": extrem}
print(f"Extremler: {len(extrem)} at (sık: Arap<={SIK_ARAP}g/İng<={SIK_ING}g, "
      f"uzun: Arap>={UZUN_ARAP}g/İng>={UZUN_ING}g) | hedef: {out['extremler']['hedef']}")

# Şehir -> TJK program linki (AGF düğmesi il-duyarlı açılır) — varsa göm
try:
    out["sehir_link"] = json.load(open("sehir_link.json", encoding="utf-8"))
    _il_say = len(out["sehir_link"].get("program", out["sehir_link"]))
    _agf_var = "agf1" in out["sehir_link"]
    print(f"Şehir linkleri: {_il_say} il gömüldü" + (" (AGF linkli)" if _agf_var else ""))
except Exception:
    out["sehir_link"] = {}

# BABA ADLARI: Sayfa1 program sayfasından (Mesafe orjin tablosunda gösterilir)
# Anahtar = normalize koşu başlığı (il ayracı gerektirmez, başlık zaten benzersiz)
out["babalar"] = {}
try:
    if "Sayfa1" in wb.sheetnames:
        def _normb(s):
            return re.sub(r"\s+", " ", str(s or "").strip()).upper()
        _cur = None
        for _row in wb["Sayfa1"].iter_rows(values_only=True):
            _c0 = str(_row[0] or "").strip()
            if re.match(r"^\d+\.\s*Koşu", _c0):
                _cur = _normb(_c0)
                out["babalar"].setdefault(_cur, {})
                continue
            if _cur and _c0.isdigit() and len(_row) > 3:
                _orj = str(_row[3] or "")
                _baba = re.sub(r"\(.*?\)", "", _orj.split("-")[0]).strip()
                if _baba:
                    out["babalar"][_cur][_c0] = _baba
        print(f"Baba adları: {sum(len(v) for v in out['babalar'].values())} at, "
              f"{len(out['babalar'])} koşu")
except Exception as _e:
    print(f"  NOT: baba adları okunamadı: {_e}")

# Koşu saatleri (tjk_yeni_yer.py üretir) — her koşu başlığına Saat alanı ekle
try:
    _saatler = json.load(open("saatler.json", encoding="utf-8"))
    _n = 0
    for _b in out.get("Sayfa1", []):
        _il = _b["header"].get("İl", "")
        _kno = str(_b["header"].get("Koşu No", "")).strip()
        _s = _saatler.get(_il, {}).get(_kno, "")
        if _s:
            _b["header"]["Saat"] = _s
            _n += 1
    out["saatler"] = _saatler
    print(f"Koşu saatleri: {_n} koşuya saat eklendi")
except Exception:
    out["saatler"] = {}

# Geç çıkış raporu (gec_cikis_rapor.py üretir) — varsa siteye göm
try:
    out["gec_cikis"] = json.load(open("gec_cikis.json", encoding="utf-8"))
    print(f"Geç çıkış: {sum(1 for v in out['gec_cikis'].values() if v.get('problem'))} problemli at gömüldü")
except Exception:
    out["gec_cikis"] = {}

# DERECELER sekmesi: 'derece' sayfası (at bazında geçmiş koşular + stil üçgeni)
out["dereceler"] = {}
if "derece" in wb.sheetnames:
    ws = wb["derece"]
    drows = list(ws.iter_rows(values_only=True))
    if drows:
        hdr = {str(h).strip(): i for i, h in enumerate(drows[0]) if h}
        def _g(r, ad):
            i = hdr.get(ad)
            v = r[i] if (i is not None and i < len(r)) else None
            return "" if v is None else str(v)
        for r in drows[1:]:
            at = _g(r, "At Adı").strip().upper()
            if not at:
                continue
            out["dereceler"].setdefault(at, []).append([
                _g(r, "Tarih"), _g(r, "Şehir"), _g(r, "Pist"), _g(r, "Pist Durumu"),
                _g(r, "Mesafe"), _g(r, "Derece"), _g(r, "Koşu Cinsi"), _g(r, "Sıra"),
                _g(r, "HP"), _g(r, "Stil"), _g(r, "Stil Üçgen"),
            ])
        print(f"Dereceler: {len(out['dereceler'])} at, "
              f"{sum(len(v) for v in out['dereceler'].values())} geçmiş koşu")

json.dump(out, open("mockup_parsed.json", "w", encoding="utf-8"), ensure_ascii=False)
