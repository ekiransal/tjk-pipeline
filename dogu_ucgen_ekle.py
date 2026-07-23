#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""DOGU UCGEN EKLEYICI - dogu_ucgen.json -> SONUC ucgen kolonlari.
Sadece 3 dogu ili, sadece BOS hucreler. Zincirden her aksam cagrilir."""
import json, re
from openpyxl import load_workbook

DOGU = {"ELAZIĞ", "DİYARBAKIR", "ŞANLIURFA"}
TARIH_RE = re.compile(r"^\d{2}\.\d{2}\.\d{4}$")
URL_RE = re.compile(r"/videoftp/\d{4}/\d{1,2}/(\d+)\.mp4\s*$")
UCGEN_STR = {1: "▷▷▷▶", 2: "▷▷▶▷", 3: "▷▶▷▷"}

def tr_up(s):
    return str(s or "").replace("i", "İ").replace("ı", "I").upper().strip()

d = json.load(open("dogu_ucgen.json", encoding="utf-8"))
# harita: (tarih, sehirUP, kosu, atno) -> 1/2/3
har = {}
for anahtar, v in d.items():
    if not v.get("serit"):
        continue
    tarih, sehir, kosu = anahtar.split("|")
    for atno, u in (v.get("ucgen") or {}).items():
        har[(tarih, sehir, int(kosu), str(atno))] = int(u)
print("  ucgen haritasi: %d at-kaydi (%d kosudan)" % (len(har), sum(1 for v in d.values() if v.get("serit"))))

# kod haritasi (link dosya adindan kosu cikarmak icin) - arsivden turet
lut = json.load(open("video_lut_arsiv.json", encoding="utf-8"))
kodlar = {}
for k, vv in lut.items():
    p = k.split("|")
    if len(p) != 3 or p[2] not in DOGU:
        continue
    url = vv[0] if isinstance(vv, (list, tuple)) else str(vv)
    m = URL_RE.search(url)
    if not m or not TARIH_RE.match(p[1]):
        continue
    import datetime
    onek = datetime.datetime.strptime(p[1], "%d.%m.%Y").strftime("%y%m%d")
    dosya = m.group(1)
    if dosya.startswith(onek):
        suf = dosya[len(onek):]
        for kl in (1, 2):
            if len(suf) > kl and suf[kl:].isdigit() and 1 <= int(suf[kl:]) <= 20:
                kodlar.setdefault(p[2], {}).setdefault(suf[:kl], 0)
                kodlar[p[2]][suf[:kl]] += 1
                break
kod = {s: max(c, key=c.get) for s, c in kodlar.items()}
print("  sehir kodlari:", kod)

def kosu_no(url, sehir, tarih):
    m = URL_RE.search(url or "")
    if not m or sehir not in kod:
        return None
    import datetime
    onek = datetime.datetime.strptime(tarih, "%d.%m.%Y").strftime("%y%m%d")
    dosya = m.group(1)
    kd = kod[sehir]
    if dosya.startswith(onek) and dosya[len(onek):].startswith(kd):
        kalan = dosya[len(onek) + len(kd):]
        return int(kalan) if kalan.isdigit() else None
    return None

wb = load_workbook("yeni_yer_SONUC.xlsx")
UC = re.compile(r"^[▷▶]{4}$")
toplam_yazilan = 0
for ad, trc, shc, ucc in (("yapılacak yer", 9, 10, 35), ("yapılacak yer 800", 3, 4, 27)):
    if ad not in wb.sheetnames:
        continue
    ws = wb[ad]
    bati_once = dogu_once = 0
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row):
        try:
            v = str(row[ucc - 1].value or "").strip()
        except IndexError:
            continue
        if UC.match(v):
            sh = tr_up(row[shc - 1].value) if len(row) >= shc else ""
            if sh in DOGU: dogu_once += 1
            else: bati_once += 1
    yazilan = 0
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row):
        try:
            tarih = str(row[trc - 1].value or "").strip()
            sehir = tr_up(row[shc - 1].value)
        except IndexError:
            continue
        if sehir not in DOGU or not TARIH_RE.match(tarih):
            continue
        huc = ws.cell(row=row[0].row, column=ucc)
        if str(huc.value or "").strip():
            continue                     # dolu hucreye ASLA dokunma
        link = str(row[42].value or "") if len(row) >= 43 else ""
        atno = str(row[43].value or "").strip() if len(row) >= 44 else ""
        atno = re.sub(r"\.0+$", "", atno)
        kn = kosu_no(link, sehir, tarih)
        if kn is None or not atno:
            continue
        u = har.get((tarih, sehir, kn, atno))
        if u:
            huc.value = UCGEN_STR[u]
            yazilan += 1
    # bati sayaci degisti mi? (yazim sadece dogu+bos hucre - degismemeli)
    bati_sonra = 0
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row):
        try:
            v = str(row[ucc - 1].value or "").strip()
        except IndexError:
            continue
        if UC.match(v):
            sh = tr_up(row[shc - 1].value) if len(row) >= shc else ""
            if sh not in DOGU: bati_sonra += 1
    assert bati_sonra == bati_once, "BATI UCGEN SAYISI DEGISTI - IPTAL"
    print("  %-20s dogu ucgen: %d vardi -> %d yazildi | bati: %d (degismedi)" % (
        ad, dogu_once, yazilan, bati_once))
    toplam_yazilan += yazilan
if toplam_yazilan == 0:
    print("  YAZILACAK UCGEN YOK - kaydetmeden cikiliyor (sayfada dogu gecmisi olmayabilir)")
    raise SystemExit(0)   # zincirde hata sayilmaz
wb.save("yeni_yer_SONUC.xlsx")
print("  KAYDEDILDI: toplam %d dogu ucgeni" % toplam_yazilan)
print("DOGU_UCGEN_EKLE_BITTI")

