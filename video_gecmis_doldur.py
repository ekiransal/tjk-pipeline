#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""VIDEO GECMIS DOLDUR - sayfadaki linksiz gecmis kosulari TJK gunluk sonuc
CSV'lerinden tamamlar. Video linki TARANMAZ, KURULUR:
    https://video-cdn.tjk.org/videoftp/{yil}/{ay}/{YYMMDD}{sehirkodu}{kosuno}.mp4
Sehir kodlari elde tutulmaz; mevcut arsivden (video_lut_arsiv.json) CIKARILIR ve
CSV ile CAPRAZ DOGRULANIR. Dogrulamayi gecemeyen sehir icin TEK LINK bile kurulmaz
(kirik link basmaktansa bos birakir). Arsiv yedeklenir, kayit SILINMEZ, sadece eklenir.

Kullanim:  python3 video_gecmis_doldur.py [max_gun_sehir]   (varsayilan 40)
"""
import json, os, re, sys, time, datetime, urllib.request, urllib.parse

KOK = os.path.dirname(os.path.abspath(__file__))
ARSIV = os.path.join(KOK, "video_lut_arsiv.json")
DURUM = os.path.join(KOK, "video_doldur_durum.json")   # bulunamayan gunler (7 gun denenmez)
HEDEF = os.path.join(KOK, "yeni_yer_SONUC.xlsx")
TARIH_RE = re.compile(r"^\d{2}\.\d{2}\.\d{4}$")
URL_RE = re.compile(r"/videoftp/(\d{4})/(\d{1,2})/(\d+)\.mp4\s*$")
SEHIRLER = ["İstanbul", "Ankara", "İzmir", "Bursa", "Adana", "Kocaeli",
            "Elazığ", "Şanlıurfa", "Diyarbakır", "Antalya", "Kayseri", "Malatya"]
KALIP = "https://video-cdn.tjk.org/videoftp/{y}/{a}/{d}.mp4"
BEKLE_SN = 1.0


def tr_up(s):
    return str(s or "").replace("i", "İ").replace("ı", "I").upper().strip()


def at_norm(s):
    t = str(s or "").strip()
    t = re.sub(r"\d+$", "", t)
    t = re.sub(r"\([^)]*\)", " ", t)
    t = re.sub(r"\s+", " ", t).strip()
    return tr_up(t)


SEHIR_AD = {tr_up(s): s for s in SEHIRLER}   # 'KOCAELİ' -> 'Kocaeli' (CSV adi)

# TJK CSV'de at adinin sonuna eklenen ekipman kisaltmalari (KG=gozluk, DB=dil bagi...)
SON_EK = {"KG", "DB", "SK", "K", "G"}


def ek_temizle(s):
    """'AKGÜL HANIM KG DB SK' -> 'AKGÜL HANIM' (ad tek kelimeye dusurulmez)."""
    par = str(s or "").split()
    while len(par) > 1 and par[-1] in SON_EK:
        par.pop()
    return " ".join(par)


def csv_url(t, sehir):
    dosya = "{0}-{1}-GunlukYarisSonuclari-TR.csv".format(t.strftime("%d.%m.%Y"), sehir)
    return "https://medya-cdn.tjk.org/raporftp/TJKPDF/{0}/{1}/CSV/GunlukYarisSonuclari/{2}".format(
        t.strftime("%Y"), t.strftime("%Y-%m-%d"), urllib.parse.quote(dosya))


def indir(url):
    try:
        r = urllib.request.urlopen(
            urllib.request.Request(url, headers={"User-Agent": "Mozilla/5.0"}), timeout=20)
        return r.read().decode("utf-8", "ignore")
    except Exception:
        return None


def csv_kosular(metin):
    """CSV metni -> {kosu_no(int): [(at_no_str, at_norm_ad), ...]}"""
    out = {}
    kosu = None
    basladi = False
    for satir in metin.splitlines():
        m = re.match(r"\s*(\d+)\.\s*Kosu", satir)
        if m:
            kosu = int(m.group(1)); basladi = False; out[kosu] = []
            continue
        if kosu is None:
            continue
        p = satir.split(";")
        if len(p) > 1 and p[0].strip() == "At No":
            basladi = True
            continue
        if basladi and len(p) >= 13 and p[0].strip().isdigit() and p[1].strip():
            out[kosu].append((p[0].strip(), ek_temizle(at_norm(p[1]))))
    return {k: v for k, v in out.items() if v}


def link_kur(tarih_d, kod, kosu_no):
    return KALIP.format(y=tarih_d.strftime("%Y"), a=str(tarih_d.month),
                        d=tarih_d.strftime("%y%m%d") + kod + str(kosu_no))


def kod_cikar(lut):
    """arsiv url'lerinden sehir -> kod. Tutarsiz sehir -> haric."""
    ham = {}   # sehirUP -> {kod_uzunlugu: {kod: sayi}}
    topl = {}  # sehirUP -> kayit sayisi
    for k, v in lut.items():
        p = k.split("|")
        if len(p) != 3:
            continue
        _, tarih, sehir = p
        url = v[0] if isinstance(v, (list, tuple)) else str(v)
        m = URL_RE.search(url)
        if not m or not TARIH_RE.match(tarih):
            continue
        d = datetime.datetime.strptime(tarih, "%d.%m.%Y")
        dosya = m.group(3)
        onek = d.strftime("%y%m%d")
        if not dosya.startswith(onek):
            continue
        suf = dosya[len(onek):]
        topl[sehir] = topl.get(sehir, 0) + 1
        for kl in (1, 2):
            if len(suf) > kl and suf[kl:].isdigit() and 1 <= int(suf[kl:]) <= 20:
                ham.setdefault(sehir, {}).setdefault(kl, {})
                ham[sehir][kl][suf[:kl]] = ham[sehir][kl].get(suf[:kl], 0) + 1
    kodlar = {}
    for sehir, kls in ham.items():
        n = topl.get(sehir, 0)
        secim = None
        for kl in (1, 2):   # kisa kod tercih; tam kapsama sart
            d = kls.get(kl) or {}
            if d:
                kod, sayi = max(d.items(), key=lambda x: x[1])
                if sayi >= max(3, int(0.95 * n)):
                    secim = kod
                    break
        if secim is not None:
            kodlar[sehir] = secim
    return kodlar, topl


def dogrula(lut, kodlar, indir_f=indir, max_gun=4):
    """Her sehir icin arsivdeki MP4-formatli gunlerin CSV'siyle capraz test:
    kurulan link == arsivdeki gercek link ? Gecemeyen sehir devre disi.
    NOT: arsivde mp4 disinda formatlar da var (YarisVideoAt sayfa linki);
    test gunleri SADECE mp4 kayitli gunlerden secilir - yoksa 0/0 olur."""
    gecen = {}
    rapor = []
    for sehir, kod in sorted(kodlar.items()):
        ad = SEHIR_AD.get(sehir)
        if not ad:
            rapor.append("  %-12s ATLANDI (CSV sehir adi bilinmiyor)" % sehir)
            continue
        gunler = sorted({k.split("|")[1] for k, v in lut.items()
                         if k.split("|")[2] == sehir
                         and URL_RE.search(v[0] if isinstance(v, (list, tuple)) else str(v))},
                        key=lambda t: datetime.datetime.strptime(t, "%d.%m.%Y"))[-max_gun:]
        uyan = toplam = 0
        for tarih in gunler:
            d = datetime.datetime.strptime(tarih, "%d.%m.%Y")
            metin = indir_f(csv_url(d, ad))
            if not metin:
                continue
            kosular = csv_kosular(metin)
            at_kosu = {}
            for kn, atlar in kosular.items():
                for no, adn in atlar:
                    at_kosu[adn] = (kn, no)
            for k, v in lut.items():
                p = k.split("|")
                if len(p) != 3 or p[1] != tarih or p[2] != sehir:
                    continue
                eş = at_kosu.get(ek_temizle(p[0]))
                if not eş:
                    continue
                url = v[0] if isinstance(v, (list, tuple)) else str(v)
                if not URL_RE.search(url):
                    continue
                toplam += 1
                if link_kur(d, kod, eş[0]) == url.strip():
                    uyan += 1
            time.sleep(BEKLE_SN)
        ok = (toplam >= 3 and uyan >= 0.95 * toplam) or (toplam == 2 and uyan == 2)
        rapor.append("  %-12s kod=%s test: %d/%d uydu -> %s" % (
            sehir, kod, uyan, toplam, "GECTI" if ok else "GECEMEDI (devre disi)"))
        if ok:
            gecen[sehir] = kod
    return gecen, rapor


def url_var(url):
    try:
        r = urllib.request.urlopen(
            urllib.request.Request(url, headers={"User-Agent": "Mozilla/5.0"}, method="HEAD"),
            timeout=12)
        return 200 <= r.status < 300
    except Exception:
        return False


def kod_kesfet(sehir, haric, indir_f=indir, urlvar_f=url_var, gun_aday=None):
    """Arsivde hic kaydi olmayan sehir (ör. dogu illeri) icin kodu KESFEDER:
    o sehrin CSV'si olan 2 gun bulunur; her aday kod icin sinir testi yapilir:
      kosu1 VAR + son kosu VAR + (son kosu+1) YOK.
    Iki gunun kesisiminde TEK kod kalirsa o kod; yoksa None (yazilmaz)."""
    ad = SEHIR_AD.get(sehir)
    if not ad:
        return None
    gunler = []
    if gun_aday is None:
        t = datetime.date.today() - datetime.timedelta(days=1)
        gun_aday = [t - datetime.timedelta(days=i) for i in range(0, 45)]
    icerikler = {}
    for g in gun_aday:
        if len(gunler) >= 2:
            break
        d = datetime.datetime(g.year, g.month, g.day)
        m = indir_f(csv_url(d, ad))
        time.sleep(BEKLE_SN)
        if m:
            k = csv_kosular(m)
            if len(k) >= 3:
                gunler.append(d)
                icerikler[d] = k
    if len(gunler) < 2:
        return None
    adaylar = [str(i) for i in list(range(10)) + list(range(10, 16)) if str(i) not in haric]
    kalan = None
    for d in gunler:
        nmax = max(icerikler[d])
        gecen = set()
        for kd in adaylar:
            if (urlvar_f(link_kur(d, kd, 1))
                    and urlvar_f(link_kur(d, kd, nmax))
                    and not urlvar_f(link_kur(d, kd, nmax + 1))):
                gecen.add(kd)
            time.sleep(0.3)
        kalan = gecen if kalan is None else (kalan & gecen)
        if not kalan:
            return None
    return kalan.pop() if kalan and len(kalan) == 1 else None


def eksik_ciftler(lut):
    """SONUC.xlsx detay satirlarindan lut'ta olmayan (tarih, sehirUP) ciftleri.
    En cok satir kapatacak cift once gelir."""
    from openpyxl import load_workbook
    if not os.path.exists(HEDEF):
        return []
    wb = load_workbook(HEDEF, read_only=True, data_only=True)
    say = {}
    for ad, atc, trc, shc in (("yapılacak yer", 2, 9, 10), ("yapılacak yer 800", 2, 3, 4)):
        if ad not in wb.sheetnames:
            continue
        for row in wb[ad].iter_rows(min_row=1, values_only=True):
            try:
                at, tarih, sehir = row[atc - 1], str(row[trc - 1] or "").strip(), row[shc - 1]
            except Exception:
                continue
            if not at or not TARIH_RE.match(tarih):
                continue
            k = "%s|%s|%s" % (at_norm(at), tarih, tr_up(sehir))
            if k not in lut:
                cift = (tarih, tr_up(sehir))
                say[cift] = say.get(cift, 0) + 1
    wb.close()
    return [c for c, _ in sorted(say.items(), key=lambda x: -x[1])]


def main():
    maxg = int(sys.argv[1]) if len(sys.argv) > 1 else 40
    if not os.path.exists(ARSIV):
        print("HATA: video_lut_arsiv.json yok"); return
    lut = json.load(open(ARSIV, encoding="utf-8"))
    print("[1] arsiv: %d kayit" % len(lut))

    kodlar, topl = kod_cikar(lut)
    print("[2] sehir kodu cikarilan: %s" % (", ".join("%s=%s" % (s, k) for s, k in sorted(kodlar.items())) or "YOK"))

    gecen, rapor = dogrula(lut, kodlar)
    print("[3] CAPRAZ DOGRULAMA (kurulan == arsivdeki gercek link?):")
    for r in rapor:
        print(r)
    if not gecen:
        print("HIC SEHIR DOGRULAMAYI GECEMEDI - hicbir sey yazilmadi, cikiliyor.")
        return

    durum = {}
    if os.path.exists(DURUM):
        try:
            durum = json.load(open(DURUM, encoding="utf-8"))
        except Exception:
            durum = {}
    bugun = datetime.date.today().isoformat()

    ciftler = eksik_ciftler(lut)
    print("[4] sayfada eksik gun-sehir cifti: %d (bu turda en fazla %d denenecek)" % (len(ciftler), maxg))

    # [4b] arsivden kodu cikarilamayan sehirler (ör. dogu illeri) icin KOD KESFI
    gereken = {s for _, s in ciftler if s in SEHIR_AD}
    kesif_gerek = sorted(gereken - set(gecen))
    haric = set(gecen.values())
    for sehir in kesif_gerek[:6]:   # tur basina en fazla 6 sehir kesfi
        # kesif gunleri: o sehrin LINKSIZ satirlarindaki tarihler (sezon disi
        # sehirler icin sart - ör. Sanliurfa yazin kosmaz, eski gunler denenir)
        gunlist = sorted({datetime.datetime.strptime(t, "%d.%m.%Y").date()
                          for t, s in ciftler if s == sehir}, reverse=True)
        kd = kod_kesfet(sehir, haric, gun_aday=(gunlist or None))
        if kd:
            print("    KOD KESFI: %s = %s (iki gunde sinir testinden gecti)" % (sehir, kd))
            gecen[sehir] = kd
            haric.add(kd)
        else:
            print("    KOD KESFI: %s bulunamadi (bu tur atlaniyor - link YAZILMADI)" % sehir)

    eklenen = 0
    denenen = 0
    for tarih, sehir in ciftler:
        if denenen >= maxg:
            break
        if sehir not in gecen or sehir not in SEHIR_AD:
            continue
        dk = "%s|%s" % (tarih, sehir)
        if dk in durum:
            try:
                if (datetime.date.today() - datetime.date.fromisoformat(durum[dk])).days < 7:
                    continue
            except Exception:
                pass
        d = datetime.datetime.strptime(tarih, "%d.%m.%Y")
        denenen += 1
        metin = indir(csv_url(d, SEHIR_AD[sehir]))
        time.sleep(BEKLE_SN)
        if not metin:
            durum[dk] = bugun
            continue
        kosular = csv_kosular(metin)
        if not kosular:
            durum[dk] = bugun
            continue
        for kn, atlar in kosular.items():
            url = link_kur(d, gecen[sehir], kn)
            for no, adn in atlar:
                k = "%s|%s|%s" % (adn, tarih, sehir)
                if k not in lut:
                    lut[k] = [url, no]
                    eklenen += 1
    print("[5] denenen gun-sehir: %d | eklenen kayit: %d" % (denenen, eklenen))

    if eklenen:
        yedek = ARSIV + ".bak_doldur_" + datetime.date.today().strftime("%Y%m%d")
        if not os.path.exists(yedek):
            json.dump(json.load(open(ARSIV, encoding="utf-8")),
                      open(yedek, "w", encoding="utf-8"), ensure_ascii=False)
        json.dump(lut, open(ARSIV, "w", encoding="utf-8"), ensure_ascii=False)
        print("[6] arsiv kaydedildi: %d kayit (yedek: %s)" % (len(lut), os.path.basename(yedek)))
    json.dump(durum, open(DURUM, "w", encoding="utf-8"), ensure_ascii=False)
    print("VIDEO_GECMIS_DOLDUR_BITTI")


if __name__ == "__main__":
    main()

