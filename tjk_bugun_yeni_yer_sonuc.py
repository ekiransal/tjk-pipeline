#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
BUGUN_SAYFA1_SAYFA2.xlsx dosyasindaki Sayfa1/Sayfa2 ile
yeni_yer_SONUC.xlsx dosyasini yeniden uretir.

Scrape yapmaz. Sadece bugun cekilmis Sayfa1/Sayfa2'yi mevcut derece, 800,
medyan, galop ve orjin referanslariyla hesaplar.
"""

import os
import shutil
from datetime import datetime

from openpyxl import load_workbook

import tjk_yeni_yer as APP
import tjk_donustur as DON
import yeni_yer_hesapla as YY


BUGUN_DOSYA = "BUGUN_SAYFA1_SAYFA2.xlsx"
CIKTI_DOSYA = "yeni_yer_SONUC.xlsx"
ILK_SEHIR = "Ankara"


def _rows(path, sheet_name):
    wb = load_workbook(path, data_only=True, read_only=True)
    if sheet_name not in wb.sheetnames:
        wb.close()
        raise SystemExit(f"{path} icinde {sheet_name} sayfasi yok.")
    rows = [list(r) for r in wb[sheet_name].iter_rows(values_only=True)]
    wb.close()
    return rows


def _data_rows(path, sheet_name):
    rows = _rows(path, sheet_name)
    if not rows:
        return []
    return rows[1:]


def _sehir_parcalari(sayfa1_rows):
    """Sayfa1'i '--- Sehir ---' ayiraclarina gore parcalar."""
    parcalar = []
    sehir = ILK_SEHIR
    aktif = []

    for row in sayfa1_rows:
        a = "" if not row else str(row[0] or "").strip()
        if a.startswith("---") and a.endswith("---"):
            if aktif:
                parcalar.append((sehir, aktif))
            sehir = a.strip("- ").strip() or sehir
            aktif = []
            continue
        aktif.append(row)

    if aktif:
        parcalar.append((sehir, aktif))
    return parcalar


def main():
    if not os.path.exists(BUGUN_DOSYA):
        raise SystemExit(f"Bugun Sayfa1/Sayfa2 dosyasi yok: {BUGUN_DOSYA}")

    sayfa1_rows = _rows(BUGUN_DOSYA, "Sayfa1")
    parcalar = _sehir_parcalari(sayfa1_rows)
    if not parcalar:
        raise SystemExit("Sayfa1 sehir parcasina ayrilamadi.")
    print(f"Bugun kaynak: Sayfa1={len(sayfa1_rows)} satir | sehirler={[s for s, _ in parcalar]}")

    print("Referanslar hazirlaniyor...")
    r800_ref, rder_ref, rmed, _ = APP.referanslari_oku()
    rder = APP.derece_al(rder_ref)
    r800 = APP.sekizyuz_al(r800_ref)
    print(f"Referans: derece={len(rder)} | 800={len(r800)} | medyan={len(rmed)}")

    refs4 = yon4 = None
    if APP.ORJIN_PANEL_REF and os.path.exists(APP.ORJIN_PANEL_REF):
        refs4, yon4 = APP.OP4.load_refs(APP.ORJIN_PANEL_REF)
        print(f"Orjin/dede referansi yuklendi: {APP.ORJIN_PANEL_REF}")

    sayfa1_all = []
    sayfa2_all = []
    yeniyer_all = []
    grids = []

    for sehir, sayfa1_city in parcalar:
        print(f"\n=== {sehir} ===")
        sayfa2_city = DON.sayfa1_to_sayfa2_rows(sayfa1_city)
        der_es, s2_at, _ = APP._kaynak_eslesme_say(sayfa2_city, rder)
        sekiz_es, _, _ = APP._kaynak_eslesme_say(sayfa2_city, r800)
        print(f"Kaynak eslesme: derece {der_es}/{s2_at} at | 800 {sekiz_es}/{s2_at} at")
        if der_es == 0 or sekiz_es == 0:
            raise SystemExit(f"HATA: {sehir} Sayfa2 ile derece/800 kaynaklari eslesmiyor; sonuc yazilmadi.")

        yeniyer_city = YY.hesapla_yeni_yer(sayfa2_city, r800, rder, rmed)
        print(f"Sayfa2={len(sayfa2_city)} | yeni yer={len(yeniyer_city)}")

        grid_city = APP.yapilacak_yer_uret(
            yeniyer_city,
            sayfa1_city,
            sehir=sehir,
            refs4=refs4,
            yon4=yon4,
        )
        print(f"yapilacak yer hucre ({sehir}): {len(grid_city)}")
        grids.append((sehir, grid_city))

        if sayfa1_all:
            sayfa1_all.append([f"--- {sehir} ---"] + [None] * 11)
        sayfa1_all.extend(sayfa1_city)
        sayfa2_all.extend(sayfa2_city)
        yeniyer_all.extend(yeniyer_city)

    yy_grid = APP._gridleri_birlestir(grids)
    print(f"\nyapilacak yer hucre toplam: {len(yy_grid)}")

    if os.path.exists(CIKTI_DOSYA):
        ts = datetime.now().strftime("%Y%m%d_%H%M%S")
        backup = f"{CIKTI_DOSYA}.onceki_{ts}.bak.xlsx"
        shutil.copy2(CIKTI_DOSYA, backup)
        print(f"Yedek alindi: {backup}")

    APP.ciktiyi_yaz(
        sayfa1_all,
        sayfa2_all,
        yeniyer_all,
        rder,
        yy_grid,
        sekizyuz_rows=r800,
        yy800_grid=None,
    )
    print(f"TAMAM: {os.path.abspath(CIKTI_DOSYA)}")


if __name__ == "__main__":
    main()
