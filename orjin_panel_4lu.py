#!/usr/bin/env python3
# -*- coding: utf-8 -*-

import argparse
import re
import shutil
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side


NA = "#YOK"
OUT_HEADERS = ["Kosu_No", "At_No", "At_Adi_Temiz", "Orjin", "Kosu_Basligi"]


TAIL_CODES = {
    "KG", "KUL", "DB", "BB", "ÇABA", "GKR", "SKG", "SKUL", "SGKR",
    "YP", "TGK", "ÖG", "DS", "SK", "K", "T",
}


def clean_horse_name(value):
    s = "" if value is None else str(value).strip()
    s = re.sub(r"\(Satılık\)", "", s, flags=re.IGNORECASE).replace("%", " ")
    s = re.sub(r"\s+", " ", s).strip()
    if not s:
        return ""
    kept = []
    for token in s.split(" "):
        if re.search(r"[a-zçğıiöşü0-9]", token):
            break
        kept.append(token)
    if kept:
        s = " ".join(kept)
    parts = s.split(" ")
    while parts:
        token = parts[-1].strip().upper()
        token = token.strip(".,;:()[]{}%*#+!?/\\")
        if len(token) <= 1 or token in TAIL_CODES:
            parts.pop()
        else:
            break
    return " ".join(parts).strip()


def trim_text(value):
    if value is None:
        return ""
    s = str(value).strip()
    trans = str.maketrans({"İ": "i", "I": "ı", "Ç": "ç", "Ğ": "ğ", "Ö": "ö", "Ş": "ş", "Ü": "ü"})
    return s.translate(trans).lower()


def key(value):
    if value is None:
        return ""
    return str(value).strip().upper()


def num(value):
    if value is None or isinstance(value, bool):
        return None
    if isinstance(value, (int, float)):
        return float(value)
    t = str(value).strip()
    if not t or t.upper() in ("#YOK", "#N/A", "#REF!", "#SAYI/0!"):
        return None
    t = t.replace("\xa0", "").replace(" ", "")
    if "." in t and "," in t:
        t = t.replace(".", "").replace(",", ".")
    elif "," in t:
        t = t.replace(",", ".")
    try:
        return float(t)
    except Exception:
        return None


def is_race_title(value):
    s = "" if value is None else str(value).strip()
    return "koşu:" in s.lower() or re.match(r"^\d{1,2}\. Koşu", s) is not None


def race_no_from_title(title):
    s = (title or "").strip()
    p = s.find(".")
    return s[:p].strip() if p > 0 else ""


def is_horse_row(ws, row):
    no = ws.cell(row, 1).value
    name = ws.cell(row, 2).value
    if str("" if no is None else no).strip() == "":
        return False
    if str("" if name is None else name).strip() == "":
        return False
    if str(name).strip().upper() == "AT İSMİ":
        return False
    try:
        n = int(round(float(str(no).replace(",", "."))))
    except Exception:
        return False
    return 1 <= n <= 99


def sire_from_origin(value):
    s = "" if value is None else str(value).strip()
    s = s.replace("–", "-").replace("—", "-")
    p = s.find("-")
    return s[:p].strip() if p > 0 else s


def dede_from_origin(value):
    s = "" if value is None else str(value).strip()
    s = s.replace("–", "-").replace("—", "-")
    p = s.find("/")
    return s[p + 1:].strip() if p >= 0 else ""


def build_orjin_rows(ws, mode):
    last_row = 0
    for r in range(ws.max_row, 0, -1):
        if any(ws.cell(r, c).value not in (None, "") for c in range(1, ws.max_column + 1)):
            last_row = r
            break
    rows = []
    race_title = ""
    race_no = ""
    r = 1
    while r <= last_row:
        a = ws.cell(r, 1).value
        b = ws.cell(r, 2).value
        if is_race_title(a):
            race_title = str(a).strip()
            race_no = race_no_from_title(race_title)
            r += 1
        elif str(a) == "N" and str(b) == "At İsmi":
            r += 1
            while r <= last_row:
                a2 = ws.cell(r, 1).value
                b2 = ws.cell(r, 2).value
                if is_race_title(a2) or (str(a2) == "N" and str(b2) == "At İsmi"):
                    break
                if is_horse_row(ws, r):
                    no = int(float(str(ws.cell(r, 1).value).replace(",", ".")))
                    name = clean_horse_name(ws.cell(r, 2).value)
                    origin = ws.cell(r, 4).value
                    root = sire_from_origin(origin) if mode == "baba" else dede_from_origin(origin)
                    rn = int(race_no) if str(race_no).isdigit() else race_no
                    rows.append([rn, no, name, root, race_title])
                r += 1
        else:
            r += 1
    return rows


def read_rows(wb, sheet):
    ws = wb[sheet]
    return [list(row) for row in ws.iter_rows(values_only=True)]


def index_rows(rows, col):
    out = {}
    for row in rows[1:]:
        if col - 1 < len(row):
            k = key(row[col - 1])
            if k and k not in out:
                out[k] = row
    return out


def val(idx, name, col):
    row = idx.get(key(name))
    if row is None or col - 1 >= len(row):
        return NA
    v = row[col - 1]
    return NA if v is None else v


def val_loose(idx, name, col):
    direct = val(idx, name, col)
    if direct != NA:
        return direct
    k = key(name)
    for rk, row in idx.items():
        if rk.startswith(k + "  (") or rk.startswith(k + " ("):
            return row[col - 1] if col - 1 < len(row) and row[col - 1] is not None else NA
    return NA


def build_analysis_rows(orjin_rows, refs):
    ku = refs["ku"]
    kalite = refs["kalite"]
    sprinter = refs["sprinter"]
    kacak = refs["kacak"]
    rows = []
    for src in orjin_rows:
        race_no, at_no, at_name, origin, title = src
        g = [None] * 40
        def setg(c, v):
            g[c - 1] = v

        setg(1, title)
        setg(2, at_no)
        setg(3, at_name)
        setg(4, origin)

        # Orta mesafe SAYI: hazır kesir sütunu yok; pay/payda'dan kur (kısa/uzun ile
        # aynı payda = yüzey toplamı). Örn: kum_Orta / kum_kum_Toplam -> '1/14'.
        def kesir(pay_col, payda_col):
            p = val(ku, origin, pay_col)
            q = val(ku, origin, payda_col)
            if p in (NA, None) or q in (NA, None):
                return NA
            try:
                return f"{int(round(float(p)))}/{int(round(float(q)))}"
            except Exception:
                return NA

        # KALİTE: Toplam Ağırlıklı Puan + sayı
        setg(7, val(kalite, origin, 2))
        setg(8, val(kalite, origin, 3))

        # MESAFE: seçilen mesafe/yüzey için oran + sayı (ORAN/SAYI yan yana)
        setg(10, val(ku, origin, 2))    # KISA KUM ORAN
        setg(11, val(ku, origin, 3))    # KISA KUM SAYI
        setg(12, val(ku, origin, 4))    # KISA ÇİM ORAN
        setg(13, val(ku, origin, 5))    # KISA ÇİM SAYI
        setg(14, val(ku, origin, 44))   # kısa sentetik oran
        setg(15, val(ku, origin, 28))   # sentetik_Kısa adet
        setg(16, val(ku, origin, 47))   # orta kum oran
        setg(17, kesir(25, 27))         # orta kum SAYI = kum_Orta / kum_kum_Toplam
        setg(18, val(ku, origin, 50))   # orta çim oran
        setg(19, kesir(21, 23))         # orta çim SAYI = çim_Orta / çim_çim_Toplam
        setg(20, val(ku, origin, 45))   # orta sentetik oran
        setg(21, kesir(29, 31))         # orta sentetik SAYI = sentetik_Orta / sentetik_Toplam
        setg(22, val(ku, origin, 8))    # UZUN KUM ORAN
        setg(23, val(ku, origin, 9))    # UZUN KUM SAYI
        setg(24, val(ku, origin, 10))   # UZUN ÇİM ORAN
        setg(25, val(ku, origin, 11))   # UZUN ÇİM SAYI
        setg(26, val(ku, origin, 46))   # uzun sentetik oran
        setg(27, val(ku, origin, 30))   # sentetik_Uzun adet

        # SPRİNTER: KARIŞIK ORAN + KARIŞIK SAYI
        setg(30, val_loose(sprinter, origin, 2))
        setg(31, val_loose(sprinter, origin, 3))

        # KAÇAK: Genel Kaçak % + Genel Kaçak Sayı
        setg(33, val(kacak, origin, 3))
        setg(34, val(kacak, origin, 4))

        # GENEL kısa/uzun (sentetik MESAFE bloğu bunları kullanır)
        setg(35, val(ku, origin, 6))    # KISA (GENEL) ORAN
        setg(36, val(ku, origin, 7))    # KISA GENEL SAYI
        setg(37, val(ku, origin, 12))   # UZUN (GENEL) ORAN
        setg(38, val(ku, origin, 13))   # UZUN GENEL SAYI
        rows.append(g)
    return rows


def parse_race_no(s):
    m = re.search(r"(\d+)\s*\.\s*Koşu", str(s or ""), flags=re.IGNORECASE)
    return int(m.group(1)) if m else 0


def parse_mesafe(s):
    for token in re.split(r"[()\-\s]+", str(s or "")):
        if token.isdigit():
            n = int(token)
            if 800 <= n <= 4000:
                return n
    return 0


def parse_zemin(s):
    t = trim_text(s)
    if "kum" in t:
        return "Kum"
    if "çim" in t or "cim" in t:
        return "Çim"
    if "sentetik" in t:
        return "Sentetik"
    return ""


def parse_irk(s):
    t = trim_text(s)
    if "arap" in t:
        return "Arap"
    if "ingiliz" in t:
        return "İngiliz"
    return ""


def lookup_yon(yon_rows, sehir, zemin, irk, mesafe):
    for row in yon_rows[1:]:
        if len(row) < 6:
            continue
        mv = num(row[3])
        if (trim_text(row[0]) == trim_text(sehir)
                and trim_text(row[1]) == trim_text(zemin)
                and trim_text(row[2]) == trim_text(irk)
                and mv is not None and int(mv) == mesafe):
            return row[4], "" if row[5] is None else str(row[5])
    return "", ""


def yorum_tipi(yorum, final):
    t = trim_text(yorum)
    if t == trim_text("Kaçak/Stalker"):
        return "KACAK", yorum
    if t == trim_text("Sprinter/Stalker"):
        return "SPRINTER", yorum
    if t in (trim_text("Nötr"), trim_text("KNSZ"), trim_text("Kazançsız")):
        return "NOTR", yorum
    f = num(final)
    if f is None:
        return "", yorum
    if f < -10:
        return "SPRINTER", "Sprinter/Stalker"
    if f > 10:
        return "KACAK", "Kaçak/Stalker"
    return "NOTR", "KNSZ"


def ana_mesafe_tipi(irk, mesafe):
    if trim_text(irk) == trim_text("Arap"):
        if mesafe <= 1500:
            return "kısa"
        if 1600 <= mesafe <= 1800:
            return "orta"
        if mesafe >= 1900:
            return "uzun"
    if trim_text(irk) == trim_text("İngiliz"):
        if mesafe <= 1400:
            return "kısa"
        if 1500 <= mesafe <= 1700:
            return "orta"
        if mesafe >= 1800:
            return "uzun"
    return ""


def pick_mesafe_cols(tip, zemin):
    z = trim_text(zemin)
    t = trim_text(tip)
    # ORTA mesafe -> MESAFE bloğu HİÇ gelmez (sadece kısa ve uzun).
    if t == trim_text("orta"):
        return (0, 0)
    return {
        ("kısa", "kum"): (10, 11),
        ("kısa", "çim"): (12, 13),
        ("kısa", "sentetik"): (35, 36),   # SENTETİK -> KISA (GENEL)
        ("uzun", "kum"): (22, 23),
        ("uzun", "çim"): (24, 25),
        ("uzun", "sentetik"): (37, 38),   # SENTETİK -> UZUN (GENEL)
    }.get((t, z), (0, 0))


def sorted_pairs(rows, r_start, r_end, no_col, val_col, ascending, count_col=0):
    pairs = []
    for i in range(r_start, r_end + 1):
        v = num(rows[i][val_col - 1]) if val_col - 1 < len(rows[i]) else None
        if v is not None:
            count = rows[i][count_col - 1] if count_col and count_col - 1 < len(rows[i]) else None
            pairs.append((rows[i][no_col - 1], v, count))
    for i in range(len(pairs) - 1):
        for j in range(i + 1, len(pairs)):
            if (ascending and pairs[j][1] < pairs[i][1]) or (not ascending and pairs[j][1] > pairs[i][1]):
                pairs[i], pairs[j] = pairs[j], pairs[i]
    return pairs


def write_block(ws, rows, r_start, r_end, top, col, title, no_col, val_col, ascending,
                count_col=0):
    pairs = sorted_pairs(rows, r_start, r_end, no_col, val_col, ascending, count_col)
    if not pairs:
        return col
    width = 4
    ws.cell(top, col).value = title
    ws.merge_cells(start_row=top, start_column=col, end_row=top, end_column=col + width - 1)
    ws.cell(top + 1, col).value = "Sıra"
    ws.cell(top + 1, col + 1).value = "No"
    ws.cell(top + 1, col + 2).value = "Değer"
    ws.cell(top + 1, col + 3).value = "Sayı"
    for i, (no_value, value, count) in enumerate(pairs, 1):
        rr = top + 1 + i
        ws.cell(rr, col).value = i
        ws.cell(rr, col + 1).value = no_value
        ws.cell(rr, col + 2).value = value
        ws.cell(rr, col + 3).value = count
    return col + width + 1


def write_panel(ws, rows, yon_rows, sehir, label):
    ws.delete_rows(1, ws.max_row)
    ws.cell(1, 1).value = f"KOŞU ANALİZ - {label.upper()}"
    ws.cell(1, 1).font = Font(bold=True, size=14)
    out_row = 3
    r_start = 0
    while r_start < len(rows):
        title = str(rows[r_start][0] or "")
        if not title:
            r_start += 1
            continue
        r_end = r_start
        while r_end + 1 < len(rows) and str(rows[r_end + 1][0] or "") == title:
            r_end += 1
        kosu_no = parse_race_no(title)
        mesafe = parse_mesafe(title)
        zemin = parse_zemin(title)
        irk = parse_irk(title)
        final_val, yorum = lookup_yon(yon_rows, sehir, zemin, irk, mesafe)
        yt, yorum = yorum_tipi(yorum, final_val)
        ana = ana_mesafe_tipi(irk, mesafe)

        is_dede = "DEDE" in str(label).upper()

        if not is_dede:
            cells = [
                ("Koşu No", kosu_no), ("İl", sehir), ("Mesafe", mesafe),
                ("Zemin", zemin), ("Irk", irk), ("Yorum", yorum), ("Final", num(final_val)),
            ]
            c = 1
            for name, value in cells:
                ws.cell(out_row, c).value = name
                ws.cell(out_row, c + 1).value = value
                c += 2

        block_top = out_row + 2
        col = 1
        col = write_block(ws, rows, r_start, r_end, block_top, col, "KALİTE", 2, 7, False, 8, olcek=0.01)
        mc, ms = pick_mesafe_cols(ana, zemin)
        if mc:
            col = write_block(ws, rows, r_start, r_end, block_top, col, "MESAFE", 2, mc, False, ms)
        # KAÇAK/SPRİNTER: DEDE panelinde ASLA gösterme. BABA'da her zaman göster
        # (yön belirsiz olsa da ikisi de gelir); KAÇAK sıralama yönü yorum tipine göre.
        if not is_dede:
            if yt == "SPRINTER":
                col = write_block(ws, rows, r_start, r_end, block_top, col, "SPRİNTER", 2, 30, False, 31)
            elif yt == "KACAK":
                col = write_block(ws, rows, r_start, r_end, block_top, col, "KAÇAK", 2, 33, False, 34)
            else:
                col = write_block(ws, rows, r_start, r_end, block_top, col, "SPRİNTER", 2, 30, False, 31)
                col = write_block(ws, rows, r_start, r_end, block_top, col, "KAÇAK", 2, 33, False, 34)

        out_row = block_top + (r_end - r_start + 1) + 6
        r_start = r_end + 1

    format_sheet(ws)


def format_sheet(ws):
    thin = Side(style="thin", color="808080")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    title_fill = PatternFill("solid", fgColor="D9E1F2")
    block_fill = PatternFill("solid", fgColor="F2F2F2")
    sub_fill = PatternFill("solid", fgColor="DDEBF7")
    summary_fill = PatternFill("solid", fgColor="FFF2CC")
    for row in ws.iter_rows():
        for cell in row:
            if cell.value is None:
                continue
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.font = Font(name="Calibri", size=10, bold=cell.row in (1,))
            cell.border = border
            if cell.row == 1:
                cell.fill = title_fill
            elif cell.value == "GENEL ÖZET":
                cell.fill = summary_fill
                cell.font = Font(name="Calibri", size=10, bold=True)
            elif cell.value in ("KALİTE", "MESAFE", "SPRİNTER", "KAÇAK"):
                cell.fill = block_fill
                cell.font = Font(name="Calibri", size=10, bold=True)
            elif cell.value in ("Sıra", "No", "Değer", "Sayı"):
                cell.fill = sub_fill
                cell.font = Font(name="Calibri", size=10, bold=True)
    for col in range(1, min(ws.max_column, 140) + 1):
        ws.column_dimensions[ws.cell(1, col).column_letter].width = 11
    ws.freeze_panes = "A3"


def ensure_sheet(wb, name):
    if name in wb.sheetnames:
        return wb[name]
    return wb.create_sheet(name)


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("input", type=Path)
    parser.add_argument("output", type=Path)
    parser.add_argument("--sehir", default="İzmir")
    args = parser.parse_args()

    args.output.parent.mkdir(parents=True, exist_ok=True)
    shutil.copyfile(args.input, args.output)

    data_wb = load_workbook(args.input, data_only=True)
    sayfa1_wb = load_workbook(args.input, data_only=False)
    out_wb = load_workbook(args.output)

    refs = {
        "ku": index_rows(read_rows(data_wb, "kısa uzun"), 1),
        "kalite": index_rows(read_rows(data_wb, "KALİTE"), 1),
        "sprinter": index_rows(read_rows(data_wb, "siprinter"), 1),
        "kacak": index_rows(read_rows(data_wb, "KAÇAK"), 1),
    }
    yon_rows = read_rows(data_wb, "YÖN")
    sayfa1 = sayfa1_wb["Sayfa1"]

    baba_rows = build_orjin_rows(sayfa1, "baba")
    dede_rows = build_orjin_rows(sayfa1, "dede")
    baba_analysis = build_analysis_rows(baba_rows, refs)
    dede_analysis = build_analysis_rows(dede_rows, refs)

    write_panel(ensure_sheet(out_wb, "YAPILACAK YER"), baba_analysis, yon_rows, args.sehir, "BABA")
    write_panel(ensure_sheet(out_wb, "YAPILACAK YER DEDE"), dede_analysis, yon_rows, args.sehir, "DEDE")

    out_wb.save(args.output)
    print(f"BABA satır: {len(baba_rows)}")
    print(f"DEDE satır: {len(dede_rows)}")
    print(f"Çıktı: {args.output}")


if __name__ == "__main__":
    main()


# =====================================================================
# PIPELINE ENTEGRASYONU (grid üretici, hızlı/iter_rows tabanlı)
# =====================================================================

def _is_horse_vals(no, name):
    if str("" if no is None else no).strip() == "" or str("" if name is None else name).strip() == "":
        return False
    if str(name).strip().upper() == "AT İSMİ":
        return False
    try:
        n = int(round(float(str(no).replace(",", "."))))
    except Exception:
        return False
    return 1 <= n <= 99


def build_orjin_rows_from_rows(rows, mode):
    """Sayfa1 satır listesi (list[list]) -> orjin satırları. ws yerine listeyle çalışır (hızlı)."""
    out = []
    title = ""
    rno = ""
    i = 0
    n = len(rows)
    while i < n:
        row = rows[i] if rows[i] is not None else []
        a = row[0] if len(row) > 0 else None
        b = row[1] if len(row) > 1 else None
        if is_race_title(a):
            title = str(a).strip()
            rno = race_no_from_title(title)
            i += 1
        elif str(a) == "N" and str(b) == "At İsmi":
            i += 1
            while i < n:
                r2 = rows[i] if rows[i] is not None else []
                a2 = r2[0] if len(r2) > 0 else None
                b2 = r2[1] if len(r2) > 1 else None
                if is_race_title(a2) or (str(a2) == "N" and str(b2) == "At İsmi"):
                    break
                if _is_horse_vals(a2, b2):
                    no = int(float(str(a2).replace(",", ".")))
                    name = clean_horse_name(b2)
                    origin = r2[3] if len(r2) > 3 else None
                    root = sire_from_origin(origin) if mode == "baba" else dede_from_origin(origin)
                    rn = int(rno) if str(rno).isdigit() else rno
                    out.append([rn, no, name, root, title])
                i += 1
        else:
            i += 1
    return out


def load_refs(ref_path):
    """Referans dosyadan (kısa uzun, KALİTE, siprinter, KAÇAK, YÖN) refs + yon döndürür."""
    wb = load_workbook(ref_path, data_only=True, read_only=True)

    def rr(sn):
        return [list(r) for r in wb[sn].iter_rows(values_only=True)]

    refs = {
        "ku": index_rows(rr("kısa uzun"), 1),
        "kalite": index_rows(rr("KALİTE"), 1),
        "sprinter": index_rows(rr("siprinter"), 1),
        "kacak": index_rows(rr("KAÇAK"), 1),
    }
    yon = rr("YÖN")
    wb.close()
    return refs, yon


def _block_grid(G, rows, r_start, r_end, top, col, title, no_col, val_col, ascending,
                count_col=0, olcek=1.0):
    """Sıra sütunu YOK: No | Değer | Sayı. Değer = round(value*olcek, 2).
    KALİTE için olcek=0.01 (120 -> 1.20)."""
    pairs = sorted_pairs(rows, r_start, r_end, no_col, val_col, ascending, count_col)
    if not pairs:
        return col
    width = 3
    G[(top, col)] = title
    G[(top + 1, col)] = "No"
    G[(top + 1, col + 1)] = "Değer"
    G[(top + 1, col + 2)] = "Sayı"
    for i, (no_value, value, count) in enumerate(pairs, 1):
        rr = top + 1 + i
        try:
            deg = round(float(value) * olcek, 2)
        except Exception:
            deg = value
        G[(rr, col)] = no_value
        G[(rr, col + 1)] = deg
        G[(rr, col + 2)] = count
    return col + width + 1


def _panel_grid(rows, yon_rows, sehir, label, top_row=1):
    """analysis satırları -> grid dict {(satır,sütun): değer}. top_row: dikey offset (çok-il)."""
    G = {}
    is_dede = "DEDE" in str(label).upper()
    G[(top_row, 1)] = f"KOŞU ANALİZ - {label.upper()} - {sehir.upper()}"
    out_row = top_row + 2
    r_start = 0
    while r_start < len(rows):
        title = str(rows[r_start][0] or "")
        if not title:
            r_start += 1
            continue
        r_end = r_start
        while r_end + 1 < len(rows) and str(rows[r_end + 1][0] or "") == title:
            r_end += 1
        kosu_no = parse_race_no(title)
        mesafe = parse_mesafe(title)
        zemin = parse_zemin(title)
        irk = parse_irk(title)
        final_val, yorum = lookup_yon(yon_rows, sehir, zemin, irk, mesafe)
        yt, yorum = yorum_tipi(yorum, final_val)
        ana = ana_mesafe_tipi(irk, mesafe)

        if not is_dede:
            cells = [("Koşu No", kosu_no), ("İl", sehir), ("Mesafe", mesafe),
                     ("Zemin", zemin), ("Irk", irk), ("Yorum", yorum), ("Final", num(final_val))]
            c = 1
            for name, value in cells:
                G[(out_row, c)] = name
                G[(out_row, c + 1)] = value
                c += 2

        block_top = out_row + 2
        col = 1
        col = _block_grid(G, rows, r_start, r_end, block_top, col, "KALİTE", 2, 7, False, 8, olcek=0.01)
        mc, ms = pick_mesafe_cols(ana, zemin)
        if mc:
            col = _block_grid(G, rows, r_start, r_end, block_top, col, "MESAFE", 2, mc, False, ms)
        # KAÇAK/SPRİNTER: DEDE'de ASLA; BABA'da her zaman (yön belirsiz olsa da).
        if not is_dede:
            if yt == "SPRINTER":
                col = _block_grid(G, rows, r_start, r_end, block_top, col, "SPRİNTER", 2, 30, False, 31)
            elif yt == "KACAK":
                col = _block_grid(G, rows, r_start, r_end, block_top, col, "KAÇAK", 2, 33, False, 34)
            else:
                col = _block_grid(G, rows, r_start, r_end, block_top, col, "SPRİNTER", 2, 30, False, 31)
                col = _block_grid(G, rows, r_start, r_end, block_top, col, "KAÇAK", 2, 33, False, 34)

        out_row = block_top + (r_end - r_start + 1) + 6
        r_start = r_end + 1
    return G


def hesapla_panel(sayfa1_rows, refs, yon_rows, sehir, mode, top_row=1):
    """Tek çağrı: scraped Sayfa1 + refs -> panel grid. mode 'baba'|'dede'."""
    orjin_rows = build_orjin_rows_from_rows(sayfa1_rows, mode)
    analysis = build_analysis_rows(orjin_rows, refs)
    label = "BABA" if mode == "baba" else "DEDE"
    return _panel_grid(analysis, yon_rows, sehir, label, top_row=top_row)


def panel_bloklari(sayfa1_rows, refs, yon_rows, sehir, mode):
    """Her koşu için AYRI panel bloğu döndürür: {kosu_no: grid} (1-tabanlı, başlık row1).
    yapılacak yer içine, koşu konumuna yerleştirmek için kullanılır."""
    orjin_rows = build_orjin_rows_from_rows(sayfa1_rows, mode)
    analysis = build_analysis_rows(orjin_rows, refs)
    is_dede = (mode == "dede")
    out = {}
    meta = {}   # kosu_no -> (final_sayi, yorum, yt)  -> özet satırı için
    i = 0
    n = len(analysis)
    while i < n:
        title = str(analysis[i][0] or "")
        if not title:
            i += 1
            continue
        j = i
        while j + 1 < n and str(analysis[j + 1][0] or "") == title:
            j += 1
        kno = parse_race_no(title)
        mesafe = parse_mesafe(title)
        zemin = parse_zemin(title)
        irk = parse_irk(title)
        final_val, yorum = lookup_yon(yon_rows, sehir, zemin, irk, mesafe)
        yt, yorum = yorum_tipi(yorum, final_val)
        ana = ana_mesafe_tipi(irk, mesafe)
        meta[kno] = (num(final_val), yorum, yt)

        G = {}
        # Başlık satırı (Koşu No/İl/Mesafe/Zemin/Irk/Yorum/Final) SADECE BABA'da.
        # DEDE tarafında tekrar gereksiz; doğrudan bloklarla başlar (baba ile hizalı).
        if not is_dede:
            cells = [("Koşu No", kno), ("İl", sehir), ("Mesafe", mesafe), ("Zemin", zemin),
                     ("Irk", irk), ("Yorum", yorum), ("Final", num(final_val))]
            c = 1
            for name, value in cells:
                G[(1, c)] = name
                G[(1, c + 1)] = value
                c += 2

        block_top = 3
        col = 1
        col = _block_grid(G, analysis, i, j, block_top, col, "KALİTE", 2, 7, False, 8, olcek=0.01)
        mc, ms = pick_mesafe_cols(ana, zemin)
        if mc:
            col = _block_grid(G, analysis, i, j, block_top, col, "MESAFE", 2, mc, False, ms)
        if not is_dede:
            # SPRINTER -> sadece SPRİNTER | KACAK -> sadece KAÇAK | NÖTR/belirsiz -> ikisi de
            if yt == "SPRINTER":
                col = _block_grid(G, analysis, i, j, block_top, col, "SPRİNTER", 2, 30, False, 31)
            elif yt == "KACAK":
                col = _block_grid(G, analysis, i, j, block_top, col, "KAÇAK", 2, 33, False, 34)
            else:
                col = _block_grid(G, analysis, i, j, block_top, col, "SPRİNTER", 2, 30, False, 31)
                col = _block_grid(G, analysis, i, j, block_top, col, "KAÇAK", 2, 33, False, 34)
        out[kno] = G
        i = j + 1
    return out, meta


def ozet_metni(final_sayi, yorum):
    """Panelin üstüne yazılacak '+/- tip' özeti. Örn: '-66 Sprinter/Stalker'."""
    if final_sayi is None:
        f_txt = ""
    else:
        try:
            f_txt = str(int(round(float(final_sayi))))
        except Exception:
            f_txt = str(final_sayi)
    return (f_txt + " " + str(yorum or "")).strip()
