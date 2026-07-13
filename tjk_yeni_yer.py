#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
TJK  UÇTAN UCA  ->  'yeni yer'
==============================

Tek komutla çalışır:

    python3 tjk_yeni_yer.py

Yaptığı iş (hepsi Python, makro/Excel formülü YOK):
  1) Bugünkü TJK günlük yarış programını çeker  -> Sayfa1 (ham at listesi)
  2) Sayfa1'i 10'lu Sayfa2'ye dönüştürür          (eski makronun karşılığı)
  3) Referans veritabanlarını (800, derece, medyan) REFERANS dosyasından okur
     (bu sayfalar SABİT FORMAT, DEĞİŞKEN içeriktir — sen güncellersin)
  4) Formül katmanını uygular                      -> yeni yer
  5) Tek bir .xlsx üretir: Sayfa1 + Sayfa2 + yeni yer

Aynı klasörde şu üç dosya bulunmalı:
    tjk_yeni_yer.py        (bu dosya — çalıştırılan)
    tjk_donustur.py        (Sayfa1 -> Sayfa2)
    yeni_yer_hesapla.py    (formül katmanı)

Gereksinimler:
    pip install selenium webdriver-manager pandas openpyxl beautifulsoup4 lxml

------------------------------------------------------------------------------
AYARLAR aşağıda. En önemlisi REFERANS_DOSYA: 800/derece/medyan sayfalarının
bulunduğu (senin sürekli güncellediğin) çalışma kitabı.
------------------------------------------------------------------------------
"""

import os
import re
import sys
import time

import pandas as pd
from io import StringIO
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font

# Aynı klasördeki doğrulanmış modüller
import tjk_donustur as DON
import yeni_yer_hesapla as YY
import derece_donustur as DD
import son800_donustur as S8
import yapilacak_yer as YYZ
import orjin_panel as OP
import orjin_panel_4lu as OP4
import baba_dede as BDX
import sayfa6_hesapla as S6X
try:
    import tjk_800_dominans as D8
except Exception as _d8e:
    print(f"[800 DOMİNANS] modül yüklenemedi (800 sayfası atlanır): {_d8e}")
    D8 = None


# =========================================================================
# son800 FARK ÜÇGENİ
#   fark = Son800Son - Son800Baş  (800 sayfası L - K), her zaman >= 0.
#   0 = KAÇAK (kendi girip bitirmiş).  Eşikler:
#       0      -> ▷▷▷▶  (kaçak / en önde)
#       1-40   -> ▷▷▶▷
#       41-90  -> ▷▶▷▷
#       91+    -> ▶▷▷▷  (en arkadan)
#   Dolu = ▶ (derecedeki stil üçgeniyle aynı yön). Boş/sayısal değilse "".
# =========================================================================
def _fark_ucgen(v):
    try:
        f = float(str(v).replace(",", "."))
    except Exception:
        return ""
    f = abs(f)         # fark = büyük - küçük (ilk giren - kendi), her zaman >= 0
    if f == 0:
        idx = 3
    elif f <= 40:
        idx = 2
    elif f <= 90:
        idx = 1
    else:
        idx = 0
    return "".join("▶" if i == idx else "▷" for i in range(4))


# =========================================================================
# AYARLAR
# =========================================================================

# 800 / derece / medyan referans sayfalarının bulunduğu çalışma kitabı.
# Bu sayfalar SABİT FORMAT ama içerikleri değişkendir; sen güncel tutarsın.
REFERANS_DOSYA = "buuuuuuuuuuuuuuuuuuuuuuu.xlsm"

# Referans çalışma kitabındaki sayfa adları (gerekirse değiştir)
SHEET_800     = "800"
SHEET_DERECE  = "derece"
SHEET_MEDYAN  = "800 ve derece çekilecek medyanl"

# Çıktı dosyası
CIKTI_DOSYA = "yeni_yer_SONUC.xlsx"

# 2. AŞAMA: yeni yer -> yapılacak yer üretilsin mi?
YAPILACAK_YER = True

# 800 DOMİNANS: yapılacak yer'in birebir formatında (orjin/dede/galop/son galop)
# ama dominans 375 günlük 800 referansından gelen AYRI bir "800 dominans" sayfası.
# son800 Veri'sindeki genişletilmiş kolonlar (Eski HP / Saha medyan / listeler) kullanılır.
SEKIZYUZ_DOMINANS = True

# yapılacak yer için ek girdiler (orjin/dede/galop/son galop) — ayrı üreticilerden.
# Her biri (dosya_yolu, sayfa_adi) ya da None. Dosya yoksa o blok atlanır.
YY_ORJIN     = ("girdiler.xlsx", "orjin")
YY_DEDE      = ("girdiler.xlsx", "dede")
YY_GALOP     = ("tjk_ULTRA_MESAFELI_YARIN.xlsx", "xy")
YY_SONGALOP  = ("tjk_ULTRA_MESAFELI_YARIN.xlsx", "xx")

# orjin/dede panelini OTOMATİK üret (AnalizCalistir karşılığı).
# (workbook, sehir) ver -> içindeki Sayfa5(model)+Sayfa6(veri)'den panel üretilir
# ve YY_ORJIN/YY_DEDE yerine o kullanılır. None -> yukarıdaki sayfadan okunur.
#   Örn: YY_ORJIN_GEN = ("baba.xlsm", "İzmir")
YY_ORJIN_GEN = None
YY_DEDE_GEN  = None
# Panel üreticinin okuyacağı sayfa adları (baba workbook'unda)
GEN_SAYFA5 = "Sayfa5"
GEN_SAYFA6 = "Sayfa6"

# === YENİ 4'LÜ ORJIN/DEDE PANELİ ===
# Eski orjin/dede (yapılacak yer içine gömülü) KALDIRILDI. Artık ayrı iki sayfa:
# "orjin (baba)" ve "orjin (dede)" — KALİTE/MESAFE/SPRİNTER/KAÇAK blokları.
# Referans dosya bu sayfaları içermeli: 'kısa uzun', 'KALİTE', 'siprinter', 'KAÇAK', 'YÖN'.
# Sen sadece bu dosyayı güncel tut; program (Sayfa1) taze scrape'ten gelir.
ORJIN_PANEL_REF = "YENI_ORJINLER_BABA_DEDE_PANEL_4LU.xlsx"

# (eski yöntem — artık kullanılmıyor, geriye dönük dursun)
YY_ORJIN_FULL = None
YY_DEDE_FULL  = None
GEN_CEKILECEK = "ÇEKİLECEK YER"
GEN_KOU       = "kısa orta uzun"

# Sayfa1 kaynağı:
#   "scrape" -> bugünkü programı siteden çek (varsayılan, elle yapıştırma YOK)
#   "ref"    -> REFERANS_DOSYA içindeki mevcut "Sayfa1"i kullan (scrape'i atla; test için)
SAYFA1_KAYNAGI = "scrape"

# "ref" modunda (test) tek Sayfa1 hangi ile ait sayılsın:
REF_SEHIR = "İzmir"

# 'derece' kaynağı (elle kopyala-yapıştır-say işini bitirir):
#   "ref"  -> REFERANS_DOSYA içindeki mevcut 'derece' sayfasını kullan
#   "veri" -> geçmiş scraper'ının çıktısını (DERECE_VERI_DOSYA) OKU + otomatik derece'ye çevir
#   "calistir" -> önce DERECE_SCRAPER_PY'yi çalıştır, sonra çıktısını derece'ye çevir
# "calistir" -> scraper'ı kendisi çalıştırır (tam otomatik). "veri" -> hazır çıktıyı okur.
# Scraper'ları AYRI çalıştırdığın için "veri" yapıldı (tekrar çekmesin).
DERECE_KAYNAGI = "veri"

# Geçmiş scraper'ın Stil'li çıktısı (Veri_Stil_Ekli sayfası -> derece + Stil).
# Not: video fazı bunu '..._STIL_EKLI.xlsx' olarak üretir.
DERECE_VERI_DOSYA = "TUM_ILLER_TUM_ATLAR_EKSIKSIZ_ILK4_SON6AY_ERTESI_GUN_STIL_EKLI.xlsx"

# "calistir" modunda çalıştırılacak scraper .py (aynı klasörde)
DERECE_SCRAPER_PY = "ERTESI_GUN_MODU_OKA_BASAN_PLUS_FAZ30_STIL_TEK_PY.py"

# '800' kaynağı (derece ile aynı mantık):
#   "ref"  -> REFERANS_DOSYA içindeki mevcut '800' sayfası
#   "veri" -> son800 scraper çıktısını (SEKIZYUZ_VERI_DOSYA) OKU + otomatik '800'e çevir
#   "calistir" -> önce SEKIZYUZ_SCRAPER_PY'yi çalıştır, sonra çıktısını '800'e çevir
SEKIZYUZ_KAYNAGI = "veri"
SEKIZYUZ_VERI_DOSYA = "SON800_BUGUN_ILK_AT_STIL_FINAL_TEST.xlsx"
SEKIZYUZ_SCRAPER_PY = "son800_BUGUN_ILK_AT_STIL_FINAL_TEST.py"
# OPTION B: son800 %45 AYRI video fazı çıktısı (son800_stil_ekle.py üretir).
# VARSA yapılacak yer 800'deki %45 stil buradan (öncelikli) gelir; yoksa derece %45'i.
SEKIZYUZ_STIL_DOSYA = "SON800_BUGUN_ILK_AT_STIL_FINAL_TEST_STIL_EKLI.xlsx"

# Galop/son galop üreticisini ana programdan çalıştır mı?
# Sen galop'u ayrı çalıştırdığın için False (hazır tjk_ULTRA_MESAFELI.xlsx okunur).
GALOP_URET = False
GALOP_SCRAPER_PY = "tjk_tum_akis_main_SELF_HEAL_KOSU_LINK_FIX_YARIN_TAM.py"

# Hangi şehirler çekilsin (TJK günlük programındaki Türkiye şehirleri)
SEHIRLER = [
    "İstanbul", "Bursa", "İzmir", "Kocaeli", "Ankara",
    "Adana", "Şanlıurfa", "Diyarbakır", "Elazığ", "Antalya",
]

# Kaç gün sonrasının programı çekilsin? (0=bugün, 1=yarın, 2=2 gün sonra ...)
# ARTIK MUTLAK: hedef = (gerçek bugün + GUN_OFFSET). Site hangi günü gösterirse
# göstersin (TJK bazen dünü gösterir) program o MUTLAK tarihe götürülür.
# Değeri TEK YERDEN ayarla: gun_ayar.py içindeki OFFSET (derece/800/galop ile aynı).
try:
    from gun_ayar import OFFSET as GUN_OFFSET
except Exception:
    GUN_OFFSET = 0

# Chrome arkada (görünmez) çalışsın mı?
HEADLESS = True

ANA_URL = "https://www.tjk.org/TR/yarissever/Info/Page/GunlukYarisProgrami"

# Sayfa1 at tablosu başlığı (makronun beklediği düzen)
SAYFA1_BASLIK = ["N", "At İsmi", "Yaş", "Orijin(Baba - Anne)", "Sıklet",
                 "Jokey", "Sahip", "Antrenör", "St", "HP", "Son 6 Y.", "KGS"]


# =========================================================================
# 1) BUGÜNKÜ PROGRAM SCRAPER  ->  Sayfa1 ham satırları
# =========================================================================

def _driver_baslat():
    from selenium import webdriver
    from selenium.webdriver.chrome.service import Service
    from webdriver_manager.chrome import ChromeDriverManager
    opts = webdriver.ChromeOptions()
    if HEADLESS:
        opts.add_argument("--headless=new")
        opts.add_argument("--window-size=1920,1080")
    else:
        opts.add_argument("--start-maximized")
    opts.add_argument("--disable-gpu")
    opts.add_argument("--no-sandbox")
    opts.add_argument("--disable-dev-shm-usage")
    opts.add_argument("--blink-settings=imagesEnabled=false")
    opts.page_load_strategy = "eager"
    d = webdriver.Chrome(service=Service(ChromeDriverManager().install()), options=opts)
    d.set_page_load_timeout(60)
    return d


def _sayfa_program_tarihi(driver):
    """Program başlığındaki tarihi (DD/MM/YYYY) yakalar."""
    try:
        kaynak = driver.page_source or ""
    except Exception:
        kaynak = ""
    m = re.search(r"(\d{2}/\d{2}/\d{4})\s+\w+\s*-\s*Yarış Programı", kaynak)
    return m.group(1) if m else ""


def _ileri_tikla(driver):
    """Program sayfasında 'ileri/sonraki gün' okuna bir kez tıklar. Başarılıysa True."""
    from selenium.webdriver.common.by import By
    try:
        els = driver.find_elements(By.TAG_NAME, "a") + driver.find_elements(By.TAG_NAME, "button")
    except Exception:
        els = []
    for el in els:
        try:
            txt = (el.text or "").strip()
            title = (el.get_attribute("title") or "").strip().lower()
            aria = (el.get_attribute("aria-label") or "").strip().lower()
            cls = (el.get_attribute("class") or "").strip().lower()
            if (txt in [">", "›", "»"] or "sonraki" in title or "ileri" in title or "next" in title
                    or "sonraki" in aria or "ileri" in aria or "next" in aria or "next" in cls):
                try:
                    el.click()
                except Exception:
                    driver.execute_script("arguments[0].click();", el)
                return True
        except Exception:
            continue
    try:
        return bool(driver.execute_script("""
            const els = Array.from(document.querySelectorAll('a,button'));
            const e = els.find(x => {
                const t = (x.innerText || x.textContent || '').trim();
                const title = (x.getAttribute('title') || '').toLowerCase();
                const aria = (x.getAttribute('aria-label') || '').toLowerCase();
                return t === '>' || t === '›' || t === '»' ||
                       title.includes('sonraki') || title.includes('ileri') || title.includes('next') ||
                       aria.includes('sonraki') || aria.includes('ileri') || aria.includes('next');
            });
            if (e) { e.click(); return true; }
            return false;
        """))
    except Exception:
        return False


def _hedef_gune_git(driver, hedef):
    """Programı MUTLAK hedef tarihe (DD/MM/YYYY) götürür. Site hangi günü
    gösterirse göstersin ileri tıklayarak hedefe ulaşır; zaten hedefteyse tıklamaz.
    Böylece gün ortasında site dünden bugüne kayar/kaymaz fark etmeksizin
    her zaman aynı (doğru) güne gidilir."""
    print(f"\nHEDEF GÜN (mutlak): {hedef}")
    for adim in range(0, 8):
        tarih_simdi = _sayfa_program_tarihi(driver)
        print(f"  kontrol {adim} | sayfa: {tarih_simdi or 'OKUNAMADI'} | hedef: {hedef}")
        if tarih_simdi == hedef:
            print(f"  hedefe ulaşıldı: {hedef}")
            return
        if not _ileri_tikla(driver):
            raise SystemExit(f"İleri gün oku bulunamadı; hedef {hedef} sayfasına gidilemedi.")
        for _ in range(1, 31):
            time.sleep(1)
            ts = _sayfa_program_tarihi(driver)
            if ts and ts != tarih_simdi:
                break
    raise SystemExit(f"Hedef güne ({hedef}) ulaşılamadı; site farklı tarih gösteriyor.")


def _norm_header(c):
    t = str(c).strip().upper()
    t = (t.replace("İ", "I").replace("Ş", "S").replace("Ğ", "G")
           .replace("Ü", "U").replace("Ö", "O").replace("Ç", "C"))
    return re.sub(r"\s+", "", t)


def _kolon_indeks(df, *adaylar):
    for i, c in enumerate(df.columns):
        cn = _norm_header(c)
        for a in adaylar:
            if cn == a or a in cn:
                return i
    return None


def _temiz_at_adi(x):
    """Web'deki kirli at adını temizler:
       'PUYOL  (3)  KGKapalı gözlük...'  ->  'PUYOL'
    Başlangıç numarası '(N)' ve sonrasındaki ekipman/uyarı metnini atar."""
    s = re.split(r"\s*\(\s*\d+\s*\)", str(x or "").strip(), maxsplit=1)[0]
    return re.sub(r"\s+", " ", s).strip()


def _temiz_siklet_ham(x):
    """'55+0.20Fazla Kilo' -> '55+0.20' (sonra sum_plus_weight 55.2 yapar). '57' -> '57'."""
    s = str(x or "").replace(" ", "").replace(",", ".")
    m = re.match(r"[0-9.+]+", s)
    return m.group(0) if m else ""


def _kosu_basligi(table, global_no):
    """Tablodan önceki 'N. Koşu' düğümünden itibaren koşula ait koşul metnini toplar;
    'Ikramiye' öncesinde keser ve başına global koşu numarasını ekler."""
    node = table.find_previous(string=re.compile(r"\d{1,2}\.\s*Ko[şs]u"))
    parts, cur, steps = [], node, 0
    while cur is not None and steps < 500:
        steps += 1
        nxt = cur.next_element
        if nxt is table or nxt is None:
            break
        if isinstance(nxt, str):
            t = nxt.strip()
            if t:
                parts.append(t)
        cur = nxt
    cond = " ".join(" ".join(parts).split())
    for kw in ("Ikramiye", "İkramiye", "Yetistirici", "Yetiştirici"):
        i = cond.find(kw)
        if i > 0:
            cond = cond[:i].strip(" ,")
            break
    return (f"{global_no}. Koşu {cond}").strip()


def _html_to_sayfa1(html, global_no=0):
    """Bir şehrin program HTML'inden Sayfa1 ham satırlarını üretir.
    global_no: o ana kadar işlenmiş koşu sayısı (şehirler arası benzersiz Kosu_No için)."""
    from bs4 import BeautifulSoup
    soup = BeautifulSoup(html, "html.parser")
    rows = []
    for table in soup.find_all("table"):
        try:
            # thousands=None: TR verisinde virgül ONDALIK; pandas varsayılanı onu
            # binlik ayıracı sanıp '60,5'->'605' yapıyordu (kilo bozulması).
            df = pd.read_html(StringIO(str(table)), thousands=None)[0]
        except Exception:
            continue
        cn = " ".join(str(c) for c in df.columns).upper().replace("İ", "I")
        if not (("ATISMI" in cn.replace(" ", "")) or ("AT ISMI" in cn)) or "SIKLET" not in cn:
            continue
        global_no += 1
        rows.append([_kosu_basligi(table, global_no)] + [None] * 11)
        rows.append(list(SAYFA1_BASLIK))
        idx = {
            "N":   _kolon_indeks(df, "N", "NO"),
            "AT":  _kolon_indeks(df, "ATISMI", "ATADI"),
            "YAS": _kolon_indeks(df, "YAS"),
            "ORJ": _kolon_indeks(df, "ORIJIN", "BABA"),
            "SIK": _kolon_indeks(df, "SIKLET"),
            "JOK": _kolon_indeks(df, "JOKEY"),
            "SAH": _kolon_indeks(df, "SAHIP"),
            "ANT": _kolon_indeks(df, "ANTRENOR"),
            "ST":  _kolon_indeks(df, "ST"),
            "HP":  _kolon_indeks(df, "HP"),
            "S6":  _kolon_indeks(df, "SON6Y", "SON6"),
            "KGS": _kolon_indeks(df, "KGS"),
        }
        for _, row in df.iterrows():
            row = list(row)
            def g(k):
                i = idx[k]
                if i is None or i >= len(row):
                    return None
                v = row[i]
                return None if pd.isna(v) else v
            n_val, at_val = g("N"), g("AT")
            if n_val is None or at_val is None:
                continue
            try:
                ni = int(float(str(n_val).strip().replace(",", ".")))
            except Exception:
                continue
            if ni < 1 or ni > 99:
                continue
            s1 = [None] * 12
            s1[0]  = ni
            s1[1]  = _temiz_at_adi(at_val)
            s1[2]  = g("YAS")
            s1[3]  = g("ORJ")
            s1[4]  = _temiz_siklet_ham(g("SIK"))
            s1[5]  = g("JOK")
            s1[6]  = g("SAH")
            s1[7]  = g("ANT")
            s1[8]  = g("ST")
            s1[9]  = g("HP")
            s1[10] = g("S6")
            s1[11] = g("KGS")
            rows.append(s1)
    return rows, global_no


def bugunku_programi_cek():
    """Bugünkü TJK programını çeker.
    Döndürür: list[(sehir, sayfa1_rows)]  — HER İL AYRI, koşu no kendi içinde 1..N."""
    from selenium.webdriver.common.by import By

    driver = _driver_baslat()
    sehir_sayfa1 = []
    try:
        driver.get(ANA_URL)
        time.sleep(2)
        print("AKTİF URL:", driver.current_url)

        # GÜN HİZALAMA (MUTLAK): sabit tarih (gun_ayar.HEDEF_TARIH) ya da bugün+OFFSET.
        from datetime import datetime, timedelta
        try:
            from gun_ayar import hedef_gun as _hedef_gun
            hedef_tarih = _hedef_gun().strftime("%d/%m/%Y")
        except Exception:
            hedef_tarih = (datetime.now() + timedelta(days=GUN_OFFSET)).strftime("%d/%m/%Y")
        _hedef_gune_git(driver, hedef_tarih)

        sehir_link = {}
        for a in driver.find_elements(By.TAG_NAME, "a"):
            try:
                href = a.get_attribute("href") or ""
                txt = (a.text or "").strip()
                if "SehirId=" not in href or "GunlukYarisProgrami" not in href:
                    continue
                for s in SEHIRLER:
                    if s.lower() in txt.lower() and s not in sehir_link:
                        sehir_link[s] = href
            except Exception:
                continue
        print("Bulunan şehirler:", list(sehir_link.keys()))
        if not sehir_link:
            print("UYARI: Türkiye şehri bulunamadı. Bugün Türkiye yarışı olmayabilir.")
        # WEB: AGF düğmesi il-duyarlı açılsın diye şehir->TJK link haritasını kaydet.
        # AGF popup kalıbı (Selenium ile tespit edildi):
        #   https://www.tjk.org/AGFv2/{SehirId}/{GGAAYYYY}/TR/{altiliNo}/1
        try:
            import json as _json
            _agf1, _agf2 = {}, {}
            try:
                _tk = str(hedef_tarih).replace("/", "").replace(".", "")   # GG/AA/YYYY -> GGAAYYYY
                for _s, _u in sehir_link.items():
                    _m = re.search(r"SehirId=(\d+)", _u)
                    if _m and len(_tk) == 8:
                        _sid = _m.group(1)
                        _agf1[_s] = f"https://www.tjk.org/AGFv2/{_sid}/{_tk}/TR/1/1"
                        _agf2[_s] = f"https://www.tjk.org/AGFv2/{_sid}/{_tk}/TR/2/1"
            except Exception as _e2:
                print(f"      NOT: AGF linkleri üretilemedi: {_e2}")
            if os.path.isdir("web"):
                _json.dump({"program": sehir_link, "agf1": _agf1, "agf2": _agf2},
                           open(os.path.join("web", "sehir_link.json"), "w",
                                encoding="utf-8"), ensure_ascii=False)
                print(f"      web/sehir_link.json yazıldı ({len(sehir_link)} şehir, AGF linkli)")
        except Exception as _e:
            print(f"      NOT: sehir_link.json yazılamadı: {_e}")

        saatler = {}   # {sehir: {"1":"14:30", ...}}  -> web/saatler.json
        kosu_idler = {}   # {sehir: {"1":"223765", ...}} -> web/kosu_id.json
        for sehir, url in sehir_link.items():
            print(f"\n--- {sehir} ---")
            try:
                driver.get(url)
                time.sleep(2)
                try:
                    driver.execute_script("window.scrollTo(0, document.body.scrollHeight);")
                    time.sleep(1)
                    driver.execute_script("window.scrollTo(0, 0);")
                    time.sleep(1)
                except Exception:
                    pass
                html = driver.page_source
            except Exception as e:
                print(f"  {sehir} açılamadı: {e}")
                continue
            # KOŞU SAATLERİ: program listesindeki "1. Koşu 14.30" düğümlerinden topla
            try:
                for m in re.finditer(r"(\d{1,2})\.\s*Ko[şs]u\s*(\d{1,2})[.:](\d{2})", html):
                    kno, sa, dk = m.group(1), m.group(2), m.group(3)
                    saatler.setdefault(sehir, {}).setdefault(kno, f"{sa}:{dk}")
            except Exception:
                pass
            # KOŞU KİMLİKLERİ: sekme bağlantıları href="#223765">1. Koşu ... kalıbından
            try:
                for m in re.finditer(r'href="#(\d{4,9})"[^>]*>\s*(\d{1,2})\.\s*Ko[şs]u', html):
                    kid, kno = m.group(1), m.group(2)
                    kosu_idler.setdefault(sehir, {}).setdefault(kno, kid)
            except Exception:
                pass
            satirlar, kosu_say = _html_to_sayfa1(html, 0)   # HER İL kendi içinde 1..N
            if satirlar:
                sehir_sayfa1.append((sehir, satirlar))
            print(f"  {sehir}: {kosu_say} koşu, {len(satirlar)} ham satır")

        # WEB: koşu saatleri (başlık yanında rozet olarak gösterilir)
        try:
            import json as _json2
            if os.path.isdir("web"):
                _json2.dump(saatler, open(os.path.join("web", "saatler.json"), "w",
                                          encoding="utf-8"), ensure_ascii=False)
                print(f"      web/saatler.json yazıldı ({sum(len(v) for v in saatler.values())} koşu saati)")
        except Exception as _e:
            print(f"      NOT: saatler.json yazılamadı: {_e}")
        # WEB: koşu kimlikleri (AGF çipi -> TJK tek-koşu görünümü)
        try:
            import json as _json3
            if os.path.isdir("web"):
                _json3.dump(kosu_idler, open(os.path.join("web", "kosu_id.json"), "w",
                                             encoding="utf-8"), ensure_ascii=False)
                print(f"      web/kosu_id.json yazıldı ({sum(len(v) for v in kosu_idler.values())} koşu kimliği)")
        except Exception as _e:
            print(f"      NOT: kosu_id.json yazılamadı: {_e}")
    finally:
        try:
            driver.quit()
        except Exception:
            pass

    print(f"\nToplam il: {len(sehir_sayfa1)}")
    return sehir_sayfa1


# =========================================================================
# Referans okuma
# =========================================================================

def _sheet_rows(wb, name):
    if name not in wb.sheetnames:
        raise SystemExit(f"Referans dosyada '{name}' sayfası yok: {REFERANS_DOSYA}")
    return [list(r) for r in wb[name].iter_rows(values_only=True)]


def referanslari_oku():
    if not os.path.exists(REFERANS_DOSYA):
        raise SystemExit(
            f"Referans dosya bulunamadı: {REFERANS_DOSYA}\n"
            f"800/derece/medyan sayfalarını içeren dosyayı bu klasöre koy "
            f"ya da betikteki REFERANS_DOSYA yolunu düzelt."
        )
    wb = load_workbook(REFERANS_DOSYA, data_only=True, read_only=True)
    r800 = _sheet_rows(wb, SHEET_800)                 # başlık satırı da veridir
    rder = _sheet_rows(wb, SHEET_DERECE)              # A1 de veridir
    rmed = _sheet_rows(wb, SHEET_MEDYAN)[1:]          # başlık satırı atılır
    sayfa1_ref = _sheet_rows(wb, "Sayfa1") if "Sayfa1" in wb.sheetnames else None
    wb.close()
    return r800, rder, rmed, sayfa1_ref


def derece_al(rder_ref):
    """DERECE_KAYNAGI'ya göre 'derece' satırlarını (başlıksız) döndürür."""
    if DERECE_KAYNAGI == "ref":
        return rder_ref

    if DERECE_KAYNAGI == "calistir":
        import subprocess
        if not os.path.exists(DERECE_SCRAPER_PY):
            raise SystemExit(f"derece scraper bulunamadı: {DERECE_SCRAPER_PY}")
        print(f"      derece scraper çalıştırılıyor: {DERECE_SCRAPER_PY} (uzun sürebilir)...")
        subprocess.run([sys.executable, DERECE_SCRAPER_PY], check=True)

    # "veri" veya "calistir" -> scraper çıktısını oku + derece formatına çevir.
    # Önce Stil'li dosya (_STIL_EKLI), yoksa temel dosyaya düş (Stil boş kalır).
    temel = "TUM_ILLER_TUM_ATLAR_EKSIKSIZ_ILK4_SON6AY_ERTESI_GUN.xlsx"
    aday = None
    for f in [DERECE_VERI_DOSYA, temel]:
        if os.path.exists(f):
            aday = f
            break
    if aday is None:
        raise SystemExit(
            f"derece Veri dosyası bulunamadı: {DERECE_VERI_DOSYA} (veya {temel})\n"
            f"Geçmiş scraper'ını çalıştırıp çıktısını bu klasöre koy "
            f"ya da DERECE_KAYNAGI='ref' yap."
        )
    if aday != DERECE_VERI_DOSYA:
        print(f"      NOT: Stil'li dosya yok, temel dosya kullanılıyor ({aday}) -> Stil(BY) boş kalır")
    rows = DD.veri_dosyasindan_derece(aday)
    print(f"      derece otomatik üretildi (Veri->derece): {len(rows)} satır [{aday}]")
    return rows


def sekizyuz_al(r800_ref):
    """SEKIZYUZ_KAYNAGI'ya göre '800' satırlarını (başlıksız) döndürür."""
    if SEKIZYUZ_KAYNAGI == "ref":
        return r800_ref

    if SEKIZYUZ_KAYNAGI == "calistir":
        import subprocess
        if not os.path.exists(SEKIZYUZ_SCRAPER_PY):
            raise SystemExit(f"son800 scraper bulunamadı: {SEKIZYUZ_SCRAPER_PY}")
        print(f"      son800 scraper çalıştırılıyor: {SEKIZYUZ_SCRAPER_PY} (uzun sürebilir)...")
        subprocess.run([sys.executable, SEKIZYUZ_SCRAPER_PY], check=True)

    if not os.path.exists(SEKIZYUZ_VERI_DOSYA):
        raise SystemExit(
            f"son800 Veri dosyası bulunamadı: {SEKIZYUZ_VERI_DOSYA}\n"
            f"son800 scraper'ını çalıştırıp çıktısını bu klasöre koy "
            f"ya da SEKIZYUZ_KAYNAGI='ref' yap."
        )
    rows = S8.veri_dosyasindan_800(SEKIZYUZ_VERI_DOSYA)
    print(f"      800 otomatik üretildi (son800->800): {len(rows)} satır")
    return rows


# =========================================================================
# Çıktı yazımı
# =========================================================================

def _oku_sayfa(kaynak):
    """(dosya, sayfa) -> list[list] satırlar; dosya/sayfa yoksa None."""
    if not kaynak:
        return None
    dosya, sayfa = kaynak
    if not os.path.exists(dosya):
        print(f"      NOT: yapılacak yer girdisi yok, atlanıyor: {dosya}")
        return None
    wb = load_workbook(dosya, data_only=True, read_only=True)
    if sayfa not in wb.sheetnames:
        print(f"      NOT: '{sayfa}' sayfası yok ({dosya}), atlanıyor")
        wb.close()
        return None
    rows = [list(r) for r in wb[sayfa].iter_rows(values_only=True)]
    wb.close()
    return rows


def _grid_to_rows(G):
    """{(r,c):val} -> list[list] (1-tabanlı)."""
    if not G:
        return []
    maxr = max(r for (r, c) in G)
    maxc = max(c for (r, c) in G)
    return [[G.get((r, c)) for c in range(1, maxc + 1)] for r in range(1, maxr + 1)]


def _orjin_panel_uret(gen):
    """gen=(workbook, sehir) -> orjin/dede paneli satırları (Sayfa5+Sayfa6'dan)."""
    dosya, sehir = gen
    if not os.path.exists(dosya):
        print(f"      NOT: panel kaynağı yok, atlanıyor: {dosya}")
        return None
    wb = load_workbook(dosya, data_only=True, read_only=True)
    if GEN_SAYFA5 not in wb.sheetnames or GEN_SAYFA6 not in wb.sheetnames:
        print(f"      NOT: {dosya} içinde {GEN_SAYFA5}/{GEN_SAYFA6} yok, atlanıyor")
        wb.close()
        return None
    s5 = [list(r) for r in wb[GEN_SAYFA5].iter_rows(values_only=True)]
    s6 = [list(r) for r in wb[GEN_SAYFA6].iter_rows(values_only=True)]
    wb.close()
    Gp = OP.hesapla(s5, s6, sehir)
    print(f"      orjin/dede paneli üretildi ({dosya}, {sehir})")
    return _grid_to_rows(Gp)


def _orjin_full_uret(sayfa1_rows, gen, mode):
    """Sayfa1 -> Sayfa2(mode) -> Sayfa6 -> panel. gen=(ref_workbook, il)."""
    dosya, sehir = gen
    if not os.path.exists(dosya):
        print(f"      NOT: referans DB yok, atlanıyor: {dosya}")
        return None
    wb = load_workbook(dosya, data_only=True, read_only=True)
    eksik = [s for s in (GEN_CEKILECEK, GEN_KOU, GEN_SAYFA5) if s not in wb.sheetnames]
    if eksik:
        print(f"      NOT: {dosya} içinde {eksik} yok, atlanıyor")
        wb.close()
        return None
    cy = [list(r) for r in wb[GEN_CEKILECEK].iter_rows(values_only=True)]
    kou = [list(r) for r in wb[GEN_KOU].iter_rows(values_only=True)]
    s5 = [list(r) for r in wb[GEN_SAYFA5].iter_rows(values_only=True)]
    wb.close()
    s2 = BDX.sayfa1_to_sayfa2(sayfa1_rows, mode=mode)
    s6 = S6X.hesapla_sayfa6(s2, cy, kou)
    Gp = OP.hesapla(s5, [[None] * 61] + s6, sehir)
    print(f"      {mode} paneli sıfırdan üretildi ({dosya}, {sehir}) — {len(s2)} at")
    return _grid_to_rows(Gp)


def yapilacak_yer_uret(yeniyer_rows, sayfa1_rows, sehir=None, refs4=None, yon4=None, tek_il=False):
    """yeni yer (=ana) -> yapılacak yer grid'i (G dict).
    sehir verilirse orjin/dede o ile göre üretilir (çok-illi gün için)."""
    ana_rows = [list(YY.YENI_YER_BASLIK)] + [list(r) for r in yeniyer_rows]
    G, _ = YYZ.hesapla(ana_rows)

    # YENİ 4'lü orjin (baba) + dede panellerini yapılacak yer İÇİNE, her koşunun
    # ESKİ orjin/dede konumuna (başlık satırı - 30) yerleştir: baba solda, dede sağda.
    if refs4 is not None and yon4 is not None:
        bb, meta4 = OP4.panel_bloklari(sayfa1_rows, refs4, yon4, sehir or "", "baba")
        dd, _ = OP4.panel_bloklari(sayfa1_rows, refs4, yon4, sehir or "", "dede")
        titles = [r for (r, c) in list(G.keys()) if c == 1 and YYZ._baslik_kosu_mu(G.get((r, 1)))]
        for r in sorted(set(titles)):
            kno = YYZ._baslik_kosu_no(G[(r, 1)])
            if kno <= 0:
                continue
            hedef = r - 30
            if hedef < 1:
                continue
            gb = bb.get(kno)
            gd = dd.get(kno)
            baba_w = 0
            if gb:
                for (br, bc), v in gb.items():
                    G[(hedef + br - 1, bc)] = v
                baba_w = max(c for (rr, c) in gb.keys())
            if gd:
                sc = baba_w + 2
                for (dr, dc), v in gd.items():
                    G[(hedef + dr - 1, sc + dc - 1)] = v

    galop = _oku_sayfa(YY_GALOP)
    if galop:
        YYZ.galop_yerlestir(G, galop, sehir=sehir, tek_il=tek_il)
    songalop = _oku_sayfa(YY_SONGALOP)
    if songalop:
        YYZ.songalop_yerlestir(G, songalop, sehir=sehir, tek_il=tek_il)
    return G


def ciktiyi_yaz(sayfa1_rows, sayfa2_rows, yeniyer_rows, derece_rows=None, yy_grid=None,
                sekizyuz_rows=None, yy800_grid=None, dom800_harita=None,
                stil800_harita=None):
    wb = Workbook()

    ws1 = wb.active
    ws1.title = "Sayfa1"
    for row in sayfa1_rows:
        ws1.append(["" if v is None else v for v in row])

    ws2 = wb.create_sheet("Sayfa2")
    DON.yaz_sayfa2(ws2, sayfa2_rows)

    ws3 = wb.create_sheet("yeni yer")
    ws3.append(YY.YENI_YER_BASLIK)
    for c in range(1, len(YY.YENI_YER_BASLIK) + 1):
        ws3.cell(1, c).font = Font(bold=True)
    for row in yeniyer_rows:
        ws3.append(["" if v is None else v for v in row])

    # derece otomatik üretildiyse onu da ekle (kontrol için)
    if derece_rows is not None and DERECE_KAYNAGI != "ref":
        ws4 = wb.create_sheet("derece")
        ws4.append(DD.DERECE_BASLIK)
        for c in range(1, len(DD.DERECE_BASLIK) + 1):
            ws4.cell(1, c).font = Font(bold=True)
        for row in derece_rows:
            ws4.append(["" if v is None else v for v in row])

    # 800 (son 800) referansı ayrı sayfa (kontrol için)
    if sekizyuz_rows is not None and SEKIZYUZ_KAYNAGI != "ref":
        ws6 = wb.create_sheet("800")
        ws6.append(S8.SEKIZYUZ_BASLIK)
        for c in range(1, len(S8.SEKIZYUZ_BASLIK) + 1):
            ws6.cell(1, c).font = Font(bold=True)
        for row in sekizyuz_rows:
            ws6.append(["" if v is None else v for v in row])

    # Stil sayısını görsel üçgene çeviren yardımcı (derece_donustur'dan)
    try:
        from derece_donustur import stil_ucgen as _stil_ucgen
    except Exception:
        _stil_ucgen = None

    # yapılacak yer (2. aşama) — Stil (col 53) yanına col 54 = ÜÇGEN
    if yy_grid:
        if _stil_ucgen:
            for (r, c) in [k for k in list(yy_grid.keys()) if k[1] == 53]:
                u = _stil_ucgen(yy_grid.get((r, 53)))
                if u:
                    yy_grid[(r, 54)] = u
        ws5 = wb.create_sheet("yapılacak yer")
        maxr = max((r for (r, c) in yy_grid), default=0)
        maxc = max((c for (r, c) in yy_grid), default=0)

        # İSTEK (v7): Toplam derece sayfasında 800 kısmı (kolon 1-19) SİLİNİR,
        # derece kolonları SOLA yaslanır (eski 20-54 -> yeni 1-35). 800'ler zaten
        # 'yapılacak yer 800' sayfasında. Paneller (kalite/mesafe/galop/orjin) AYNEN kalır.
        import re as _re
        _TARIH_RE = _re.compile(r"^\d{2}\.\d{2}\.\d{4}$")

        def _txt(v):
            s = "" if v is None else str(v).strip()
            return s

        def _tarih_mi(v):
            return bool(_TARIH_RE.match(_txt(v)))

        def _metin_mi(v):
            s = _txt(v)
            if s == "" or s in ("Değer", "No", "Sayı"):
                return False
            try:
                float(s.replace(",", "."))
                return False
            except Exception:
                return True

        _kaydirilan = 0
        for r in range(1, maxr + 1):
            # AT satırı (800+derece) ya da derece DEVAM satırı -> derece kısmını sola kaydır.
            # DÜZELTME (Bursa 4 vakası): galop hücreleri "1 / -2,20 / R" formatındadır ("/" içerir);
            # son galop blokları tesadüfen kolon 21'e denk gelince derece satırı sanılıp
            # 19 kolon sola kayıyordu (ilk veri satırı uçuyordu). At adında "/" olmaz.
            _c2  = _txt(yy_grid.get((r, 2)))
            _c21 = _txt(yy_grid.get((r, 21)))
            at_800 = (_metin_mi(yy_grid.get((r, 2))) and _tarih_mi(yy_grid.get((r, 3)))
                      and "/" not in _c2)
            derece_devam = (_metin_mi(yy_grid.get((r, 21))) and _tarih_mi(yy_grid.get((r, 28)))
                            and "/" not in _c21)
            if at_800 or derece_devam:
                row = [yy_grid.get((r, c), None) for c in range(20, 55)]  # eski 20-54 -> 1-35
                if any(v not in (None, "") for v in row):
                    ws5.append(row)
                    _kaydirilan += 1
                # derece kısmı tamamen boşsa (sadece 800'ü olan at) satır atlanır
            else:
                ws5.append([yy_grid.get((r, c), None) for c in range(1, maxc + 1)])
        print(f"[SÜRÜM] yapılacak yer = SOLA YASLI derece (800 kısmı silindi, v7) | kaydırılan satır: {_kaydirilan}")

    # yapılacak yer 800 (AYRI sayfa; dominans 375 günlük 800 referansından)
    #   İSTEK: AT satırlarında S sütunundan (19) sonraki DERECE kısmı silinsin,
    #   yalnızca 800 verisi (A-S) kalsın; 6'lı 800-dominans (eski 44-49) S'nin
    #   arkasına (kolon 20-25) eklensin. ÜSTTEKİ orjin/galop/koşu isimleri AYNEN kalır.
    if yy800_grid:
        try:
            from tjk_donustur import clean_horse_name as _clean
        except Exception:
            _clean = lambda s: str(s or "").strip().upper()
        harita = dom800_harita or {}
        stil_map = stil800_harita or {}

        def _at_satiri(r):
            """col2 metin (at adı) ise at satırı; sayısal/boş ise panel/galop/başlık."""
            v = yy800_grid.get((r, 2))
            if v is None or str(v).strip() == "":
                return False
            try:
                float(str(v).replace(",", "."))
                return False
            except Exception:
                return True

        print(f"[SÜRÜM] yapılacak yer 800 = PER-KOŞU dominans + %45 STİL + galop(25-42), derece devam satırları atılır (v6)")
        _yerlesen = 0
        _ucgen_say = 0
        _stil_say = 0
        _atilan_devam = 0
        _out = []   # SIRALI çıktı satırları (devam satırları atlanır, boşluk bırakmaz;
                    # koşu araları boş satırları korunur)
        for r in sorted(set(rr for (rr, _c) in yy800_grid)):
            if _at_satiri(r):
                row = [yy800_grid.get((r, c)) for c in range(1, 20)]   # 1-19 (800 verisi)
                # 800-dominans KOŞU BAZINDA (harita -> yoksa grid 44-49)
                dom6 = [None] * 6
                _kondu = False
                try:
                    _kkey = D8.kosu_key(yy800_grid.get((r, 2)), yy800_grid.get((r, 3)),
                                        yy800_grid.get((r, 4)), yy800_grid.get((r, 5)),
                                        yy800_grid.get((r, 8)))
                except Exception:
                    _kkey = None
                dom = harita.get(_kkey) if _kkey else None
                if dom:
                    for i, val in enumerate(dom):
                        if val is not None and str(val) != "":
                            dom6[i] = val
                            _kondu = True
                if not _kondu:
                    for i in range(6):
                        v = yy800_grid.get((r, 44 + i))
                        if v is not None and str(v).strip() != "":
                            dom6[i] = v
                            _kondu = True
                if _kondu:
                    _yerlesen += 1
                # %45 STİL (col26) + STİL ÜÇGENİ (col27)
                stil26 = None
                sttri27 = None
                stil = stil_map.get(_kkey) if _kkey else None
                if stil is not None and str(stil).strip() != "":
                    stil26 = stil
                    t = D8.stil_ucgen(stil)
                    if t:
                        sttri27 = t
                        _stil_say += 1
                # son800 FARK ÜÇGENİ (col28)
                ftri = _fark_ucgen(yy800_grid.get((r, 15)))
                if ftri:
                    _ucgen_say += 1
                row = row + dom6 + [stil26, sttri27, ftri]   # 19+6+3 = 28 kolon
                row = row + [None] * (42 - len(row))
                _out.append(row)
            else:
                ilk19 = any(yy800_grid.get((r, c)) not in (None, "") for c in range(1, 20))
                # DERECE DEVAM SATIRI: 1-19 boş & 20-23 dolu (koşu no + at adı col21 + kilo).
                # GALOP satırları cols 20-23 BOŞ (içerik col24+'dan başlar) -> KORUNUR.
                son2023 = any(yy800_grid.get((r, c)) not in (None, "") for c in range(20, 24))
                if (not ilk19) and son2023:
                    _atilan_devam += 1
                    continue   # derece devam satırı -> boşluk bırakmadan atla
                row = [yy800_grid.get((r, c)) for c in range(1, 43)]
                # KOŞU ARASI BOŞLUK: yeni "Koşu No" bloğuna geçerken 2 boş satır bırak.
                if isinstance(row[0], str) and row[0].strip() == "Koşu No" and _out:
                    while _out and all(v in (None, "") for v in _out[-1]):
                        _out.pop()
                    _out.append([None] * 42)
                    _out.append([None] * 42)
                _out.append(row)
        print(f"[SÜRÜM] 800 PER-KOŞU dominans yerleşen: {_yerlesen} | %45 stil üçgeni: {_stil_say} "
              f"| fark üçgeni: {_ucgen_say} | atılan derece devam satırı: {_atilan_devam}")

        ws8 = wb.create_sheet("yapılacak yer 800")
        for row in _out:
            ws8.append(row)

    # ÖNEMLİ: VLOOKUP'ta eşleşmeyenler "#N/A"/"#REF!" gibi METİN olarak gelir.
    # openpyxl bunları HATA-tipli hücre yazar ve Excel açılışta "içerik sorunu"
    # uyarısı verip bu hücreleri SİLER (derece/800 boş görünür). Hata-tipli tüm
    # hücreleri düz METNE çevir; değer görünür kalır, dosya bozulmaz.
    for ws in wb.worksheets:
        for row in ws.iter_rows():
            for cc in row:
                if cc.data_type == "e":
                    cc.data_type = "s"
                elif cc.data_type == "f":
                    # '=' ile başlayan metin (ör. '--- Kocaeli ---' veya web'den
                    # gelen bir değer) yanlışlıkla FORMÜL yazılmış olabilir; Excel
                    # bunu reddedip "içerik sorunu" verir. Düz metne çevir.
                    _val = cc.value
                    cc.data_type = "s"
                    cc._value = _val

    # KİLİDE DAYANIKLI KAYIT: yeni_yer_SONUC.xlsx Excel'de AÇIK/kilitliyse üzerine
    # yazamayabilir -> otomatik YEDEK isimlere düş, böylece gece boyu emek boşa gitmez.
    _hedefler = [CIKTI_DOSYA, "yeni_yer_SONUC_YEDEK.xlsx", "yeni_yer_SONUC_YEDEK2.xlsx"]
    _kayit = None
    for _h in _hedefler:
        try:
            wb.save(_h)
            _kayit = _h
            break
        except PermissionError:
            print(f"      ⚠ '{_h}' açık/kilitli, kaydedilemedi -> alternatif deneniyor...")
        except Exception as _e:
            print(f"      ⚠ '{_h}' kaydedilemedi ({_e}) -> alternatif deneniyor...")
    if _kayit is None:
        raise SystemExit(
            "HATA: Çıktı kaydedilemedi — yeni_yer_SONUC.xlsx ve yedekleri de açık/kilitli "
            "olabilir. Lütfen Excel'i kapatıp tjk_yeni_yer.py'yi tekrar çalıştır "
            "(sadece bu son adım; scrape/video çıktıları diskte duruyor).")
    if _kayit != CIKTI_DOSYA:
        print(f"\n⚠ '{CIKTI_DOSYA}' AÇIK olduğu için çıktı YEDEK isimle kaydedildi:\n"
              f"    {os.path.abspath(_kayit)}\n"
              f"  -> {CIKTI_DOSYA}'i kapatıp bu yedeği onun yerine koyabilirsin (hiçbir şey kaybolmadı).")
    return _kayit


# =========================================================================
# ANA AKIŞ
# =========================================================================

def _gridleri_birlestir(grids, bosluk=5):
    """Her ilin yapılacak yer grid'ini ALT ALTA birleştirir (satır offset'i ile)."""
    out = {}
    offset = 0
    for sehir, g in grids:
        if not g:
            continue
        maxr = max(r for (r, c) in g)
        for (r, c), v in g.items():
            out[(r + offset, c)] = v
        offset += maxr + bosluk
    return out


def main():
    print("=" * 60)
    print("TJK -> yeni yer + yapılacak yer  (uçtan uca, çok-illi)")
    print("=" * 60)

    # 1) Sayfa1 — HER İL AYRI
    if SAYFA1_KAYNAGI == "ref":
        print("\n[1] Sayfa1: referans dosyadan okunuyor...")
        _, _, _, sayfa1_ref = referanslari_oku()
        if not sayfa1_ref:
            raise SystemExit("Referans dosyada 'Sayfa1' yok; SAYFA1_KAYNAGI='scrape' yap.")
        il_listesi = [(REF_SEHIR, sayfa1_ref)]
    else:
        print("\n[1] Sayfa1: bugünkü program siteden çekiliyor...")
        il_listesi = bugunku_programi_cek()
        if not il_listesi:
            raise SystemExit("Program çekilemedi. İnternet/Chrome'u kontrol et.")
    print(f"      İşlenecek il: {[s for s, _ in il_listesi]}")

    # 3) Referanslar (tüm iller için ortak; bir kez)
    print("\n[2] Referans veritabanları (800/derece/medyan) hazırlanıyor...")
    r800_ref, rder_ref, rmed, _ = referanslari_oku()
    rder = derece_al(rder_ref)
    r800 = sekizyuz_al(r800_ref)
    print(f"      800: {len(r800)} (kaynak={SEKIZYUZ_KAYNAGI}) | derece: {len(rder)} (kaynak={DERECE_KAYNAGI}) | medyan: {len(rmed)}")

    # galop/son galop üreticisini çalıştır (bir kez, tüm iller)
    if YAPILACAK_YER and GALOP_URET and os.path.exists(GALOP_SCRAPER_PY):
        print("\n[3] Galop/son galop çekiliyor (uzun sürebilir)...")
        try:
            import subprocess
            subprocess.run([sys.executable, GALOP_SCRAPER_PY], check=True)
        except Exception as e:
            print(f"      UYARI: galop üreticisi hata verdi: {e}")

    # 4) HER İL: Sayfa2 -> yeni yer -> yapılacak yer
    # 4'lü orjin/dede paneli için referansları bir kez oku
    refs4 = yon4 = None
    if ORJIN_PANEL_REF and os.path.exists(ORJIN_PANEL_REF):
        try:
            refs4, yon4 = OP4.load_refs(ORJIN_PANEL_REF)
            print(f"\n[2b] Orjin/dede paneli referansı yüklendi: {ORJIN_PANEL_REF}")
        except Exception as e:
            print(f"      UYARI: orjin panel referansı okunamadı ({ORJIN_PANEL_REF}): {e}")
    elif ORJIN_PANEL_REF:
        print(f"      NOT: orjin panel referansı yok ({ORJIN_PANEL_REF}), orjin/dede sayfaları atlanacak.")

    # 800 dominans referans haritası (bir kez): son800 Veri'sinden at->eski HP/kilo+listeler
    ref_map_800 = {}
    if YAPILACAK_YER and SEKIZYUZ_DOMINANS and D8 is not None:
        try:
            ref_map_800 = D8.son800_ref_map(SEKIZYUZ_VERI_DOSYA)
        except Exception as e:
            print(f"      UYARI: 800 dominans referansı okunamadı, 800 sayfası atlanır: {e}")
            ref_map_800 = {}

    sayfa1_all, sayfa2_all, yeniyer_all = [], [], []
    grids = []
    grids_800 = []
    dom800_harita = {}   # kosu_key -> (d50hp,d50k,d66hp,d66k,d75hp,d75k)  KOŞU BAZINDA
    stil800_harita = {}  # kosu_key -> %45 seyir sırası (stil)
    for sehir, sayfa1_city in il_listesi:
        print(f"\n=== İL: {sehir} ===")
        sayfa2_city = DON.sayfa1_to_sayfa2_rows(sayfa1_city)
        yeniyer_city = YY.hesapla_yeni_yer(sayfa2_city, r800, rder, rmed)
        print(f"      Sayfa2: {len(sayfa2_city)} | yeni yer: {len(yeniyer_city)}")
        # birleşik sayfalar için biriktir (il ayraç satırı ile)
        if sayfa1_all:
            sayfa1_all.append([f"--- {sehir} ---"] + [None] * 11)
        sayfa1_all.extend(sayfa1_city)
        sayfa2_all.extend(sayfa2_city)
        yeniyer_all.extend(yeniyer_city)
        if YAPILACAK_YER:
            # orjin/dede panelleri yapılacak yer İÇİNE yerleştirilir (refs4/yon4 ile)
            _tek_il = (len(il_listesi) == 1)   # tek şehir yarışıyorsa galop fallback güvenli
            grid_city = yapilacak_yer_uret(yeniyer_city, sayfa1_city, sehir=sehir,
                                           refs4=refs4, yon4=yon4, tek_il=_tek_il)
            grids.append((sehir, grid_city))

            # 800 DOMİNANS: aynı format, dominans 375 günlük 800 referansından.
            if SEKIZYUZ_DOMINANS and D8 is not None and ref_map_800:
                yeniyer_800_city = D8.yeni_yer_800_uret(yeniyer_city, ref_map_800)
                grid_800_city = yapilacak_yer_uret(yeniyer_800_city, sayfa1_city, sehir=sehir,
                                                   refs4=refs4, yon4=yon4, tek_il=(len(il_listesi) == 1))
                grids_800.append((sehir, grid_800_city))
                # DOMİNANS DÜZELTME (BAY OLOF vakası): harita 800-DÖNÜŞTÜRÜLMÜŞ
                # satırlardan kurulur (yeniyer_800_city). Böylece eski taraf (Eski
                # HP/kilo + saha listeleri) satırın 800 KOŞUSUNA aittir ve anahtar
                # (_800 alanları) ile değer AYNI koşuyu anlatır. Eskiden NORMAL
                # satırlar kullanılıyordu: anahtar 800 koşusunu, değer derece
                # koşusunu gösteriyordu -> S8'de dominans bir satır kayıyordu.
                # STİL: %45 seyir sırası — normal satırlardan (değişmedi).
                try:
                    anaNc = [list(YY.YENI_YER_BASLIK)] + [list(r) for r in yeniyer_city]
                    ana8 = [list(YY.YENI_YER_BASLIK)] + [list(r) for r in yeniyer_800_city]
                    dom800_harita.update(D8.dominans_haritasi(ana8))
                    stil800_harita.update(D8.stil_haritasi(anaNc))
                except Exception as e:
                    print(f"      UYARI: 800 dominans/stil haritası kurulamadı: {e}")

    # OPTION B: son800 %45 (AYRI video fazı) VARSA derece %45'inin ÜSTÜNE yaz (öncelikli).
    # son800_stil_ekle.py çıktısı (SEKIZYUZ_STIL_DOSYA) yoksa hiçbir şey değişmez ->
    # yapılacak yer 800 stili derece %45'inden gelir (mevcut davranış korunur).
    if SEKIZYUZ_DOMINANS and D8 is not None:
        try:
            _s800_stil = D8.son800_stil_haritasi(SEKIZYUZ_STIL_DOSYA)
            if _s800_stil:
                stil800_harita.update(_s800_stil)
                print(f"      [STAGE2] son800 %45 stil eklendi (öncelikli): {len(_s800_stil)} koşu")
        except Exception as e:
            print(f"      NOT: son800 %45 stil eklenemedi ({e}) -> derece %45'i kullanılır")

    yy_grid = _gridleri_birlestir(grids) if YAPILACAK_YER else None
    yy800_grid = _gridleri_birlestir(grids_800) if grids_800 else None
    if yy800_grid:
        print(f"      800 dominans (tüm iller): {len(yy800_grid)} hücre")
    if yy_grid:
        maxr = max((r for (r, c) in yy_grid), default=0)
        print(f"\n      yapılacak yer (tüm iller): {len(yy_grid)} hücre | son satır: {maxr}")

    _kayitli = ciktiyi_yaz(sayfa1_all, sayfa2_all, yeniyer_all, rder, yy_grid, sekizyuz_rows=r800,
                           yy800_grid=yy800_grid, dom800_harita=dom800_harita,
                           stil800_harita=stil800_harita)
    print(f"\nTAMAM ✓  Çıktı: {os.path.abspath(_kayitli or CIKTI_DOSYA)}")


if __name__ == "__main__":
    main()
