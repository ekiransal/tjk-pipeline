#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
ORJİN KÜTÜPHANESİ -> PANEL REFERANS SAYFALARI
=============================================

Kütüphanedeki (SQLite) ham galibiyetlerden, mevcut orjin/dede panelinin
(orjin_panel_4lu) okuduğu referans sayfalarını BİREBİR düzende üretir:
  'kısa uzun', 'KALİTE', 'siprinter', 'KAÇAK'
'YÖN' ve diğer sayfalar (Sayfa1/2, YAPILACAK YER...) şablondan aynen korunur.

Böylece: ayda/yılda bir kütüphaneye yeni koşu eklersin -> bu modül panel
referansını güncel oranlarla yeniden yazar -> mevcut panel akışı hiç
değişmeden çalışır.

Panelin okuduğu kolonlar (1-tabanlı), orjin_panel_4lu.build_analysis_rows'tan:
  kısa uzun: 2,3,4,5,8,9,10,11, 21,25,28,29,30, 44,45,46,47,50
  KALİTE   : 2 (TAP), 3 (sayı "graded/total")
  siprinter: 2 (karışık oran), 3 (karışık sayı) ; anahtar "AD (İrk)" prefix
  KAÇAK    : 3 (genel %), 4 (sayı "kaçak/total")
"""

import os
import numpy as np
import pandas as pd
from openpyxl import load_workbook

import tjk_orjin_kutuphane as K


def _df(conn):
    return pd.read_sql_query("SELECT * FROM kazanan", conn)


def _baba_irk(g):
    """Babanın baskın ırkı (galibiyet çoğunluğu)."""
    vc = g["irk"].value_counts()
    return vc.index[0] if len(vc) else ""


def hesapla(conn):
    """Her baba için panel referansına lazım tüm alanları hesaplar."""
    df = _df(conn)
    df["kat"] = df.apply(lambda r: K._mesafe_tipi(r["irk"], r["mesafe"]), axis=1)
    df["puan"] = df["cins"].map(K._kalite_agirlik)
    esik = K.sprinter_esikleri(conn)

    out = {}
    for baba, g in df.groupby("baba"):
        if not str(baba).strip():
            continue
        tot = len(g)
        irk = _baba_irk(g)

        rec = {"baba": baba, "irk": irk, "toplam": tot}

        # --- MESAFE dağılımı ---
        yuzey_tot = {}
        for zem in K.YUZEYLER:
            zg = g[g["zemin"] == zem]
            yuzey_tot[zem] = len(zg)
            for kat in K.KATEGORILER:
                rec[f"{zem}_{kat}"] = int((zg["kat"] == kat).sum())
        # genel (yüzeyler toplamı) kategori sayıları
        for kat in K.KATEGORILER:
            rec[f"genel_{kat}"] = sum(rec[f"{zem}_{kat}"] for zem in K.YUZEYLER)
        rec["genel_toplam"] = sum(yuzey_tot.values())
        rec["_yuzey_tot"] = yuzey_tot

        # --- KALİTE ---
        rec["tap"] = float(g["puan"].sum())
        rec["graded"] = int((g["puan"] > 0).sum())
        rec["kalite_skoru"] = rec["tap"] / tot if tot else 0.0

        # --- KAÇAK (fark 0..4) ---
        fk = g[g["fark"].notna()]
        rec["kacak_payda"] = int(len(fk))
        kacak_mask = (fk["fark"] >= 0) & (fk["fark"] <= 4)
        rec["kacak_genel"] = int(kacak_mask.sum())
        for kat in K.KATEGORILER:
            rec[f"kacak_{kat}"] = int((kacak_mask & (fk["kat"] == kat)).sum())

        # --- SPRİNTER (fark > eşik) ---
        kar_ustu = kar_payda = 0
        for zem in K.YUZEYLER:
            zg = fk[fk["zemin"] == zem]
            payda = len(zg)
            ustu = 0
            for _, r in zg.iterrows():
                es = esik.get((r["irk"], zem))
                if es and r["fark"] > es["esik"]:
                    ustu += 1
            rec[f"spr_{zem}_ustu"] = int(ustu)
            rec[f"spr_{zem}_payda"] = int(payda)
            kar_ustu += ustu
            kar_payda += payda
        rec["spr_karisik_ustu"] = int(kar_ustu)
        rec["spr_karisik_payda"] = int(kar_payda)

        out[baba] = rec
    return out, esik


def _oran(pay, payda):
    return (pay / payda) if payda else 0.0


def _sayi(pay, payda):
    return f"{pay}/{payda}"


def yaz(conn, sablon_path, cikti_path, min_galibiyet=0):
    """Referans sayfalarını şablonun kopyasına yazar."""
    recs, esik = hesapla(conn)
    wb = load_workbook(sablon_path)

    kutuphane_isimleri = {str(b).strip().upper() for b in recs.keys()}

    def _orijinal_eksikler(sheet_ad):
        """Şablonda olup kütüphanede OLMAYAN orjinlerin orijinal satırları
        (isim, [hücreler]). Bunlar silinmeden önce yakalanır ve sona eklenir."""
        if sheet_ad not in wb.sheetnames:
            return []
        ws = wb[sheet_ad]
        korunan = []
        for row in ws.iter_rows(min_row=2, values_only=True):
            if not row or row[0] is None:
                continue
            ham = str(row[0]).strip()
            isim = ham.split("  (")[0].split(" (")[0].strip().upper()
            if isim and isim != "NAN" and isim not in kutuphane_isimleri:
                korunan.append(list(row))
        return korunan

    def temiz_yaz(sheet_ad, basliklar):
        korunan = _orijinal_eksikler(sheet_ad)
        if sheet_ad in wb.sheetnames:
            ws = wb[sheet_ad]
            ws.delete_rows(1, ws.max_row)
        else:
            ws = wb.create_sheet(sheet_ad)
        for c, b in enumerate(basliklar, 1):
            ws.cell(1, c).value = b
        ws._korunan_eksik = korunan  # sona eklemek için sakla
        return ws

    def _eksikleri_ekle(ws, r):
        """Kütüphanede olmayan (eski) orjinlerin orijinal satırlarını sona ekler."""
        for row in getattr(ws, "_korunan_eksik", []):
            for c, v in enumerate(row, 1):
                ws.cell(r, c).value = v
            r += 1
        return r

    babalar = sorted(recs.keys())

    # ---------- kısa uzun ----------
    ws = temiz_yaz("kısa uzun", ["baba"] + [""] * 49)
    ws.cell(1, 2).value = "KISA KUM ORAN"; ws.cell(1, 3).value = "KISA KUM SAYI"
    ws.cell(1, 4).value = "KISA ÇİM ORAN"; ws.cell(1, 5).value = "KISA ÇİM SAYI"
    ws.cell(1, 6).value = "KISA (GENEL) ORAN"; ws.cell(1, 7).value = "KISA GENEL SAYI"
    ws.cell(1, 8).value = "UZUN KUM ORAN"; ws.cell(1, 9).value = "UZUN KUM SAYI"
    ws.cell(1, 10).value = "UZUN ÇİM ORAN"; ws.cell(1, 11).value = "UZUN ÇİM SAYI"
    ws.cell(1, 12).value = "UZUN (GENEL) ORAN"; ws.cell(1, 13).value = "UZUN GENEL SAYI"
    r = 2
    for baba in babalar:
        d = recs[baba]
        yt = d["_yuzey_tot"]
        gtop = d["genel_toplam"]
        ws.cell(r, 1).value = baba
        ws.cell(r, 2).value = _oran(d["Kum_kısa"], yt["Kum"])
        ws.cell(r, 3).value = _sayi(d["Kum_kısa"], yt["Kum"])
        ws.cell(r, 4).value = _oran(d["Çim_kısa"], yt["Çim"])
        ws.cell(r, 5).value = _sayi(d["Çim_kısa"], yt["Çim"])
        # KISA/UZUN GENEL = çim+kum+sentetik toplamı (sentetik koşular bunu kullanır)
        ws.cell(r, 6).value = _oran(d["genel_kısa"], gtop)
        ws.cell(r, 7).value = _sayi(d["genel_kısa"], gtop)
        ws.cell(r, 8).value = _oran(d["Kum_uzun"], yt["Kum"])
        ws.cell(r, 9).value = _sayi(d["Kum_uzun"], yt["Kum"])
        ws.cell(r, 10).value = _oran(d["Çim_uzun"], yt["Çim"])
        ws.cell(r, 11).value = _sayi(d["Çim_uzun"], yt["Çim"])
        ws.cell(r, 12).value = _oran(d["genel_uzun"], gtop)
        ws.cell(r, 13).value = _sayi(d["genel_uzun"], gtop)
        # orta adetleri (panel col 21=çim_Orta, 25=kum_Orta, 28=sent_Kısa, 29=sent_Orta, 30=sent_Uzun)
        ws.cell(r, 21).value = d["Çim_orta"]
        ws.cell(r, 25).value = d["Kum_orta"]
        ws.cell(r, 28).value = d["Sentetik_kısa"]
        ws.cell(r, 29).value = d["Sentetik_orta"]
        ws.cell(r, 30).value = d["Sentetik_uzun"]
        # sentetik/orta oranları (col 44=kısa sent, 45=orta sent, 46=uzun sent, 47=orta kum, 50=orta çim)
        ws.cell(r, 44).value = _oran(d["Sentetik_kısa"], yt["Sentetik"])
        ws.cell(r, 45).value = _oran(d["Sentetik_orta"], yt["Sentetik"])
        ws.cell(r, 46).value = _oran(d["Sentetik_uzun"], yt["Sentetik"])
        ws.cell(r, 47).value = _oran(d["Kum_orta"], yt["Kum"])
        ws.cell(r, 50).value = _oran(d["Çim_orta"], yt["Çim"])
        r += 1
    _eksikleri_ekle(ws, r)

    # ---------- KALİTE ----------
    ws = temiz_yaz("KALİTE", ["baba", "Toplam Ağırlıklı Puan", "sayı", "", "", "", "", "",
                              "Toplam Kazanç", "Kalite Skoru", ""])
    r = 2
    for baba in babalar:
        d = recs[baba]
        ws.cell(r, 1).value = baba
        ws.cell(r, 2).value = round(d["tap"], 4)
        ws.cell(r, 3).value = _sayi(d["graded"], d["toplam"])
        ws.cell(r, 9).value = d["toplam"]
        ws.cell(r, 10).value = round(d["kalite_skoru"], 6)
        ws.cell(r, 11).value = float(d["graded"])
        r += 1
    _eksikleri_ekle(ws, r)

    # ---------- siprinter ----------  (anahtar 'AD  (İrk)')
    ws = temiz_yaz("siprinter", ["baba_ad", "KARIŞIK ORAN", "KARIŞIK SAYI"])
    r = 2
    for baba in babalar:
        d = recs[baba]
        ws.cell(r, 1).value = f"{baba}  ({d['irk']})" if d["irk"] else baba
        ws.cell(r, 2).value = _oran(d["spr_karisik_ustu"], d["spr_karisik_payda"])
        ws.cell(r, 3).value = _sayi(d["spr_karisik_ustu"], d["spr_karisik_payda"])
        r += 1
    _eksikleri_ekle(ws, r)

    # ---------- KAÇAK ----------
    ws = temiz_yaz("KAÇAK", ["baba", "ırk", "Genel Kaçak %", "Genel Kaçak Sayı",
                             "Genel Payda (fark dolu)", "Genel Kaçak Adet (0-4)", "",
                             "Kısa Kaçak Adet (0-4)", "Kısa Kaçak % (Genel payda)",
                             "Orta Kaçak Adet (0-4)", "Orta Kaçak % (Genel payda)",
                             "Uzun Kaçak Adet (0-4)", "Uzun Kaçak % (Genel payda)"])
    r = 2
    for baba in babalar:
        d = recs[baba]
        payda = d["kacak_payda"]
        ws.cell(r, 1).value = baba
        ws.cell(r, 2).value = d["irk"]
        ws.cell(r, 3).value = _oran(d["kacak_genel"], payda) * 100
        ws.cell(r, 4).value = _sayi(d["kacak_genel"], payda)
        ws.cell(r, 5).value = payda
        ws.cell(r, 6).value = d["kacak_genel"]
        ws.cell(r, 8).value = d["kacak_kısa"]
        ws.cell(r, 9).value = _oran(d["kacak_kısa"], payda) * 100
        ws.cell(r, 10).value = d["kacak_orta"]
        ws.cell(r, 11).value = _oran(d["kacak_orta"], payda) * 100
        ws.cell(r, 12).value = d["kacak_uzun"]
        ws.cell(r, 13).value = _oran(d["kacak_uzun"], payda) * 100
        r += 1
    _eksikleri_ekle(ws, r)

    wb.save(cikti_path)
    print(f"[PANEL] referans sayfaları yazıldı -> {cikti_path} | baba: {len(babalar)}")
    return cikti_path


if __name__ == "__main__":
    import sys
    if len(sys.argv) < 3:
        print("Kullanım: python3 tjk_orjin_panel_yaz.py <sablon_panel.xlsx> <cikti_panel.xlsx> [db]")
        raise SystemExit(1)
    conn = K.baglan(sys.argv[3] if len(sys.argv) > 3 else None)
    yaz(conn, sys.argv[1], sys.argv[2])
